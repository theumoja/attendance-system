import json
import csv
from io import BytesIO
from itertools import groupby
from operator import attrgetter

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
# Import all models from models.py
from attendance.models import *



from django.shortcuts import render
from django.contrib.auth.decorators import login_required

@login_required
def my_apps(request):
    return render(request, 'attendance/my_apps.html')

@login_required
def home(request):
    """
    Redirects the user to their appropriate dashboard based on their role.
    """
    if request.user.role == User.IS_ADMIN:
        return redirect('attendance:admin_dashboard')
    elif request.user.role == User.IS_TEACHER:
        return redirect('attendance:teacher_dashboard')
    elif request.user.role == User.IS_WARDEN:
        return redirect('attendance:warden_dashboard')
    
    elif request.user.role == User.IS_ACCOUNTANT:
        return redirect('attendance:accountant_dashboard')

    elif request.user.role == User.IS_LIBRARIAN:
        return redirect('attendance:librarian_dashboard')
        
    elif request.user.role == User.IS_STUDENT:
        return redirect('attendance:student_dashboard')

from django.db import transaction  # <-- Add this missing import
import csv
import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q


import json
import csv
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone


from django.db.models import Sum
from django.core.exceptions import PermissionDenied


from django.db.models import Sum, F

@login_required
def accountant_dashboard(request):
    """
    Accountant's central overview dashboard:
    - Student financial summary (fees billed, collected, pending)
    - Staff payroll summary (total disbursed, paid/unpaid staff)
    - Disciplinary case distribution
    - Lodging occupancy snapshot
    - Recent transaction feed
    """
    if request.user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        raise PermissionDenied("Access Denied: Only Accountants can view this dashboard.")

    current_term = AcademicTerm.objects.filter(is_current=True).first()

    # ---------- Student Fees ----------
    total_students = StudentProfile.objects.count()
    fee_accounts = StudentTermFee.objects.filter(term=current_term) if current_term else StudentTermFee.objects.none()
    total_billed = fee_accounts.aggregate(Sum('total_fees_due'))['total_fees_due__sum'] or 0
    total_collected = fee_accounts.aggregate(Sum('total_amount_paid'))['total_amount_paid__sum'] or 0
    total_pending = total_billed - total_collected
    cleared_count = fee_accounts.filter(total_fees_due__lte=F('total_amount_paid'), total_fees_due__gt=0).count()
    partial_count = fee_accounts.filter(total_amount_paid__gt=0, total_fees_due__gt=F('total_amount_paid')).count()
    unpaid_count = fee_accounts.filter(total_amount_paid=0, total_fees_due__gt=0).count()

    # ---------- Staff Payroll ----------
    total_staff = User.objects.exclude(role=User.IS_STUDENT).count()
    paid_staff = StaffPaymentRecord.objects.filter(term=current_term).values('staff').distinct().count() if current_term else 0
    unpaid_staff = total_staff - paid_staff
    total_payroll = StaffPaymentRecord.objects.filter(term=current_term).aggregate(Sum('amount'))['amount__sum'] or 0

    # ---------- Disciplinary ----------
    total_disciplinary = DisciplinaryRecord.objects.count()
    mild = DisciplinaryRecord.objects.filter(severity='MILD').count()
    severe = DisciplinaryRecord.objects.filter(severity='SEVERE').count()
    very_severe = DisciplinaryRecord.objects.filter(severity='VERY_SEVERE').count()

    # ---------- Lodging (current term) ----------
    if current_term:
        allocations = RoomAllocation.objects.filter(term=current_term)
        total_allocated = allocations.values('student').distinct().count()
        total_rooms = Room.objects.count()
        total_capacity = Room.objects.aggregate(Sum('capacity'))['capacity__sum'] or 0
        occupancy_rate = (total_allocated / total_capacity * 100) if total_capacity > 0 else 0
        # FIX: Use reg_number__in instead of id__in
        allocated_regs = allocations.values_list('student_id', flat=True)
        unallocated_students = StudentProfile.objects.exclude(reg_number__in=allocated_regs).count()
    else:
        total_allocated = 0
        total_rooms = 0
        total_capacity = 0
        occupancy_rate = 0
        unallocated_students = 0

    # ---------- Recent Transactions ----------
    recent_transactions = FeePaymentTransaction.objects.all().order_by('-date_recorded')[:5]

    context = {
        'total_students': total_students,
        'total_billed': total_billed,
        'total_collected': total_collected,
        'total_pending': total_pending,
        'cleared_count': cleared_count,
        'partial_count': partial_count,
        'unpaid_count': unpaid_count,
        'total_staff': total_staff,
        'paid_staff': paid_staff,
        'unpaid_staff': unpaid_staff,
        'total_payroll': total_payroll,
        'total_disciplinary': total_disciplinary,
        'mild': mild,
        'severe': severe,
        'very_severe': very_severe,
        'total_allocated': total_allocated,
        'total_rooms': total_rooms,
        'total_capacity': total_capacity,
        'occupancy_rate': round(occupancy_rate, 1),
        'unallocated_students': unallocated_students,
        'recent_transactions': recent_transactions,
        'current_term': current_term,
    }
    return render(request, 'attendance/accountant_dashboard.html', context)

@login_required
def warden_dashboard(request):
    """
    Warden’s central command centre:
    - Lodging statistics (hostels, rooms, occupancy)
    - Disciplinary overview
    - Personal staff payment history
    """
    if request.user.role != User.IS_WARDEN:
        raise PermissionDenied("Access Denied: Only Wardens can view this dashboard.")

    user = request.user
    current_term = AcademicTerm.objects.filter(is_current=True).first()

    # ---------- LODGING STATISTICS ----------
    hostels = Hostel.objects.all()
    total_hostels = hostels.count()
    rooms = Room.objects.all()
    total_rooms = rooms.count()
    total_capacity = rooms.aggregate(total=Sum('capacity'))['total'] or 0

    # Allocations for the current term (or all if none active)
    allocations_qs = RoomAllocation.objects.filter(term=current_term) if current_term else RoomAllocation.objects.all()
    allocated_students_count = allocations_qs.values('student').distinct().count()

    total_students = StudentProfile.objects.count()
    unallocated_students_count = total_students - allocated_students_count

    occupancy_rate = (allocated_students_count / total_capacity * 100) if total_capacity > 0 else 0

    # Per‑hostel occupancy breakdown
    hostel_occupancy = []
    for hostel in hostels:
        rooms_in_hostel = Room.objects.filter(hostel=hostel)
        capacity_sum = rooms_in_hostel.aggregate(total=Sum('capacity'))['total'] or 0
        occupied = RoomAllocation.objects.filter(
            room__hostel=hostel,
            term=current_term if current_term else None
        ).values('student').distinct().count()
        hostel_occupancy.append({
            'name': hostel.name,
            'capacity': capacity_sum,
            'occupied': occupied,
            'percentage': (occupied / capacity_sum * 100) if capacity_sum > 0 else 0
        })

    # ---------- DISCIPLINARY STATISTICS ----------
    total_records = DisciplinaryRecord.objects.count()
    mild = DisciplinaryRecord.objects.filter(severity='MILD').count()
    severe = DisciplinaryRecord.objects.filter(severity='SEVERE').count()
    very_severe = DisciplinaryRecord.objects.filter(severity='VERY_SEVERE').count()

    recent_disciplinary = DisciplinaryRecord.objects.all().order_by('-date_logged')[:5]

    # ---------- STAFF PAYMENT HISTORY ----------
    staff_payments = StaffPaymentRecord.objects.filter(staff=user).order_by('-payment_date')

    # ---------- CONTEXT ----------
    context = {
        'total_hostels': total_hostels,
        'total_rooms': total_rooms,
        'total_capacity': total_capacity,
        'total_allocated': allocated_students_count,
        'unallocated_students': unallocated_students_count,
        'occupancy_rate': round(occupancy_rate, 1),
        'hostel_occupancy': hostel_occupancy,
        'total_records': total_records,
        'mild': mild,
        'severe': severe,
        'very_severe': very_severe,
        'recent_disciplinary': recent_disciplinary,
        'staff_payments': staff_payments,
        'current_term': current_term,
    }
    return render(request, 'attendance/warden_dashboard.html', context)


from datetime import datetime, timedelta
from django.utils import timezone
from django.db import models
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from .models import Book, LibraryRecord, StudentProfile, TeacherProfile, User

@login_required
def librarian_dashboard(request):
    if request.user.role not in [User.IS_LIBRARIAN, User.IS_ADMIN]:
        return render(request, 'errors/403.html', status=403)
    
    # ---------- Statistics ----------
    total_books = Book.objects.count()
    active_loans = LibraryRecord.objects.filter(status='ISSUED').count()
    returned_books = LibraryRecord.objects.filter(status='RETURNED').count()
    
    today = timezone.now().date()
    overdue_count = LibraryRecord.objects.filter(
        status='ISSUED',
        due_date__lt=today
    ).count()
    
    # ---------- Monthly trend (last 6 months) ----------
    monthly_labels = []
    monthly_data = []
    for i in range(5, -1, -1):
        date_cursor = today - timedelta(days=30 * i)
        month_name = date_cursor.strftime('%b')
        month_num = date_cursor.month
        year_num = date_cursor.year
        count = LibraryRecord.objects.filter(
            issue_date__month=month_num, 
            issue_date__year=year_num
        ).count()
        monthly_labels.append(month_name)
        monthly_data.append(count)
    
    # ---------- Most borrowed books (top 5) ----------
    top_books = (
        LibraryRecord.objects
        .values('book__title')
        .annotate(total=models.Count('id'))
        .order_by('-total')[:5]
    )
    top_books_list = [
        {'book_title': item['book__title'], 'total': item['total']}
        for item in top_books
    ]
    
    # ---------- Recent activity (last 10) ----------
    recent_records = LibraryRecord.objects.select_related('book', 'student', 'teacher').order_by('-issue_date')[:10]
    activity_feed = []
    for rec in recent_records:
        borrower = rec.student.name if rec.student else (rec.teacher.name if rec.teacher else "Unknown")
        action = "Returned" if rec.status == 'RETURNED' else "Borrowed"
        activity_feed.append({
            'book': rec.book.title if rec.book else "Unknown Book",
            'student': borrower,
            'action': action,
            'date': rec.issue_date.strftime('%b %d, %Y'),
            'is_returned': rec.status == 'RETURNED'
        })
    
    # ---------- Overdue books with days calculated ----------
    overdue_records = LibraryRecord.objects.filter(
        status='ISSUED',
        due_date__lt=today
    ).select_related('book', 'student', 'teacher')
    
    # Attach overdue_days to each record
    for record in overdue_records:
        record.overdue_days = (today - record.due_date).days
    
    context = {
        'total_books': total_books,
        'active_loans': active_loans,
        'returned_books': returned_books,
        'overdue_count': overdue_count,
        'monthly_labels': monthly_labels,
        'monthly_data': monthly_data,
        'top_books': top_books_list,
        'activity_feed': activity_feed,
        'overdue_records': overdue_records,
        'today': today,  # <-- now available in template
    }
    return render(request, 'attendance/librarian_dashboard.html', context)

from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from .models import Book, LibraryRecord, StudentProfile, TeacherProfile, Department, ReserveRequest, User

from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone

from .models import User, LibraryRecord, ReserveRequest, Book, StudentProfile, TeacherProfile, Department

from datetime import timedelta
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
# 1. Make sure Department is imported
from .models import Book, LibraryRecord, ReserveRequest, StudentProfile, TeacherProfile, Department

def manage_library(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        # ---------- ACTION 1: Approve Reserve Request ----------
        if action == 'approve_reserve':
            reserve_id = request.POST.get('reserve_id')
            reserve_req = get_object_or_404(ReserveRequest, id=reserve_id)
            
            if reserve_req.book.available_copies < 1:
                messages.error(
                    request, 
                    f"Cannot approve request: '{reserve_req.book.title}' is out of stock."
                )
                return redirect('attendance:manage_library')

            reserve_req.status = 'APPROVED'
            reserve_req.save()

            student_borrower = getattr(reserve_req, 'student', None)
            teacher_borrower = getattr(reserve_req, 'teacher', None)
            borrower_name = student_borrower.name if student_borrower else teacher_borrower.name

            messages.success(
                request, 
                f"Approved reserve request for {borrower_name} ({reserve_req.book.title}). Ready for issuance via Quick Issue!"
            )
            return redirect('attendance:manage_library')

        # ---------- ACTION 2: Reject Reserve Request ----------
        elif action == 'reject_reserve':
            reserve_id = request.POST.get('reserve_id')
            reserve_req = get_object_or_404(ReserveRequest, id=reserve_id)
            reserve_req.status = 'REJECTED'
            reserve_req.save()

            messages.info(request, f"Reserve request for '{reserve_req.book.title}' was rejected.")
            return redirect('attendance:manage_library')

        # ---------- ACTION 3: Quick Direct Issuance ----------
        elif action == 'issue_books':
            borrower_type = request.POST.get('borrower_type')
            student_id = request.POST.get('student_id')
            teacher_id = request.POST.get('teacher_id')
            
            book_ids = request.POST.getlist('book_ids[]')
            quantities = request.POST.getlist('quantities[]')
            
            remarks = request.POST.get('remarks', '')
            due_date = timezone.now().date() + timedelta(days=14)

            student_obj = None
            teacher_obj = None
            
            if borrower_type == 'student' and student_id:
                student_obj = get_object_or_404(StudentProfile, reg_number=student_id)
            elif borrower_type == 'teacher' and teacher_id:
                teacher_obj = get_object_or_404(TeacherProfile, name=teacher_id)
                
            if not student_obj and not teacher_obj:
                messages.error(request, "Please choose a valid Student or Teacher.")
                return redirect('attendance:manage_library')
                
            if not book_ids:
                messages.error(request, "Please select at least one book to issue.")
                return redirect('attendance:manage_library')
            
            if len(book_ids) != len(quantities):
                messages.error(request, "Invalid data: book and quantity mismatch.")
                return redirect('attendance:manage_library')
                
            issued_count = 0
            for b_id, qty_str in zip(book_ids, quantities):
                try:
                    qty = int(qty_str)
                except ValueError:
                    qty = 1
                if qty < 1:
                    qty = 1
                    
                book = get_object_or_404(Book, id=b_id)
                
                # Check Reserve Collection authorization
                if book.is_reserve:
                    reserve_kwargs = {'book': book, 'status': 'APPROVED'}
                    if student_obj:
                        reserve_kwargs['student'] = student_obj
                    else:
                        reserve_kwargs['teacher'] = teacher_obj

                    approved_req = ReserveRequest.objects.filter(**reserve_kwargs).first()
                    
                    if not approved_req:
                        messages.error(
                            request, 
                            f"'{book.title}' is in the Reserve Collection and can only be issued to borrowers with an approved reserve request."
                        )
                        return redirect('attendance:manage_library')
                    
                    # Mark the reserve request as fulfilled upon issuance
                    approved_req.status = 'FULFILLED'
                    approved_req.save()

                if book.available_copies < qty:
                    messages.error(
                        request,
                        f"Not enough copies of '{book.title}'. Available: {book.available_copies}, requested: {qty}."
                    )
                    return redirect('attendance:manage_library')
                
                for _ in range(qty):
                    LibraryRecord.objects.create(
                        student=student_obj,
                        teacher=teacher_obj,
                        book=book,
                        due_date=due_date,
                        remarks=remarks,
                        status='ISSUED'
                    )
                    book.available_copies -= 1
                    book.save()
                    issued_count += 1

            messages.success(request, f"Successfully issued {issued_count} book copy/copies.")
            return redirect('attendance:manage_library')

        # ---------- ACTION 4: Return Book ----------
        elif action == 'return_book':
            record_id = request.POST.get('record_id')
            record = get_object_or_404(LibraryRecord, id=record_id)
            
            if record.status == 'RETURNED':
                messages.warning(request, f"'{record.book.title}' has already been returned.")
                return redirect('attendance:manage_library')
                
            record.status = 'RETURNED'
            record.return_date = timezone.now().date()
            record.save()
            
            # Increment stock back
            record.book.available_copies += 1
            record.book.save()
            
            # Complete/Reset associated Reserve Request
            if record.student:
                ReserveRequest.objects.filter(
                    student=record.student, 
                    book=record.book, 
                    status__in=['APPROVED', 'FULFILLED']
                ).update(status='COMPLETED')
            elif record.teacher:
                ReserveRequest.objects.filter(
                    teacher=record.teacher, 
                    book=record.book, 
                    status__in=['APPROVED', 'FULFILLED']
                ).update(status='COMPLETED')
            
            messages.success(request, f"'{record.book.title}' marked as returned.")
            return redirect('attendance:manage_library')

    # GET Request Context Setup
    books = Book.objects.all().order_by('title')
    students = StudentProfile.objects.all().order_by('name')
    teachers = TeacherProfile.objects.all().order_by('name')
    
    # 2. Query departments from database
    departments = Department.objects.all().order_by('name')

    issued_records = LibraryRecord.objects.filter(status='ISSUED').select_related('student', 'teacher', 'book').order_by('-issue_date')
    history_records = LibraryRecord.objects.filter(status='RETURNED').select_related('student', 'teacher', 'book').order_by('-return_date')[:50]
    reserve_requests = ReserveRequest.objects.select_related('student', 'teacher', 'book').order_by('-request_date')

    context = {
        'books': books,
        'students': students,
        'teachers': teachers,
        'departments': departments,  # 3. Added departments to context here
        'issued_records': issued_records,
        'history_records': history_records,
        'reserve_requests': reserve_requests,
    }
    return render(request, 'attendance/manage_library.html', context)


@login_required
def library_reader_dashboard(request):
    """
    For students and teachers:
    - Personal borrowed history
    - General catalog vs. Reserve Collection separation
    - Student & Teacher application status tracking for reserve items
    """
    if request.user.role not in [User.IS_STUDENT, User.IS_TEACHER]:
        return render(request, 'errors/403.html', status=403)

    profile = None
    borrowed_records = []
    my_reserve_requests = []

    if request.user.role == User.IS_STUDENT:
        try:
            profile = request.user.student_profile
            borrowed_records = LibraryRecord.objects.filter(student=profile).order_by('-issue_date')
            my_reserve_requests = ReserveRequest.objects.filter(student=profile).order_by('-request_date')
        except StudentProfile.DoesNotExist:
            pass
    elif request.user.role == User.IS_TEACHER:
        try:
            profile = request.user.teacher_profile
            borrowed_records = LibraryRecord.objects.filter(teacher=profile).order_by('-issue_date')
            if hasattr(ReserveRequest, 'teacher'):
                my_reserve_requests = ReserveRequest.objects.filter(teacher=profile).order_by('-request_date')
        except TeacherProfile.DoesNotExist:
            pass

    # Catalog split based on is_reserve status
    general_books = Book.objects.filter(is_reserve=False).order_by('title')
    reserve_books = list(Book.objects.filter(is_reserve=True).order_by('title'))

    # Determine user filter for LibraryRecord lookup
    rec_filter = {'student': profile} if request.user.role == User.IS_STUDENT else {'teacher': profile}

    # Map only active/blocking requests to reserve books
    active_requests_map = {}
    for req in my_reserve_requests:
        if req.book_id in active_requests_map:
            continue  # Already captured the most recent request for this book
        
        if req.status in ['PENDING', 'APPROVED']:
            active_requests_map[req.book_id] = req
        elif req.status == 'FULFILLED':
            # Only block new applications if an active LibraryRecord with status='ISSUED' still exists
            still_issued = LibraryRecord.objects.filter(
                book=req.book,
                status='ISSUED',
                **rec_filter
            ).exists()
            
            if still_issued:
                active_requests_map[req.book_id] = req

    # Attach active request (if any) to each reserve book instance
    for book in reserve_books:
        book.user_request = active_requests_map.get(book.id)

    context = {
        'borrowed_records': borrowed_records,
        'general_books': general_books,
        'reserve_books': reserve_books,
        'my_reserve_requests': my_reserve_requests,
        'profile': profile,
        'role': request.user.role,
    }
    return render(request, 'attendance/library_reader.html', context)

@login_required
@transaction.atomic
def upload_books(request):
    """
    Standalone view for librarians to add new books to the catalog.
    GET  -> display the upload form.
    POST -> process new book entry.
    """
    if request.user.role not in [User.IS_LIBRARIAN, User.IS_ADMIN]:
        return render(request, 'errors/403.html', status=403)

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        author = request.POST.get('author', '').strip()
        isbn = request.POST.get('isbn', '').strip()
        total_copies_str = request.POST.get('total_copies', '1')
        department_id = request.POST.get('department')
        
        # 1. Read the checkbox value from POST ('true', 'on', or '1')
        is_reserve = request.POST.get('is_reserve') in ['true', 'on', '1']

        if not title:
            messages.error(request, "Book title is required.")
            return redirect(request.META.get('HTTP_REFERER', 'attendance:manage_library'))

        try:
            total_copies = int(total_copies_str)
            if total_copies < 1:
                total_copies = 1
        except ValueError:
            total_copies = 1

        # Resolve department if provided
        department = None
        if department_id:
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                messages.warning(request, "Selected department not found; book will be added without a department.")

        # Check for existing book by ISBN (if provided) or exact title
        book = None
        if isbn:
            book = Book.objects.filter(isbn=isbn).first()
        if not book:
            book = Book.objects.filter(title__iexact=title).first()

        if book:
            # Update existing record
            book.total_copies += total_copies
            book.available_copies += total_copies
            if author:
                book.author = author
            if isbn:
                book.isbn = isbn
            if department:
                book.department = department
            
            # 2. Update reserve status on existing book
            book.is_reserve = is_reserve
            book.save()
            
            messages.success(
                request,
                f"Updated stock for '{book.title}'. Added {total_copies} more copy/copies."
            )
        else:
            # 3. Pass is_reserve when creating new book
            Book.objects.create(
                title=title,
                author=author,
                isbn=isbn if isbn else None,
                total_copies=total_copies,
                available_copies=total_copies,
                department=department,
                is_reserve=is_reserve
            )
            messages.success(request, f"Successfully cataloged '{title}'.")

        # Redirect back to the originating page (Manage Library or standalone upload page)
        return redirect(request.META.get('HTTP_REFERER', 'attendance:manage_library'))

    # GET – render the standalone form with department choices
    departments = Department.objects.all().order_by('name')
    return render(request, 'attendance/upload_books.html', {'departments': departments})






@login_required
@transaction.atomic
def issue_book(request):
    if request.user.role not in [User.IS_LIBRARIAN, User.IS_ADMIN]:
        return render(request, 'errors/403.html', status=403)
        
    if request.method == 'POST':
        borrower_type = request.POST.get('borrower_type')
        student_id = request.POST.get('student_id')
        teacher_id = request.POST.get('teacher_id')
        book_ids = request.POST.getlist('book_ids[]')
        quantities = request.POST.getlist('quantities[]')
        due_date_str = request.POST.get('due_date')
        remarks = request.POST.get('remarks', '')
        
        if due_date_str:
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d').date()
        else:
            due_date = timezone.now().date() + timedelta(days=14)
            
        student_obj = None
        teacher_obj = None
        
        if borrower_type == 'student':
            if not student_id:
                messages.error(request, "Select a valid Student borrower.")
                return redirect('attendance:issue_book')
            student_obj = get_object_or_404(StudentProfile, reg_number=student_id)
        elif borrower_type == 'teacher':
            if not teacher_id:
                messages.error(request, "Select a valid Teacher borrower.")
                return redirect('attendance:issue_book')
            # LOOKUP BY NAME (front-end sends the teacher's name)
            teacher_obj = get_object_or_404(TeacherProfile, name=teacher_id)
            #teacher_obj = get_object_or_404(TeacherProfile, staff_id=teacher_id)
            
        if not book_ids:
            messages.error(request, "You must select one or more books.")
            return redirect('attendance:issue_book')
        
        if len(book_ids) != len(quantities):
            messages.error(request, "Invalid data: book and quantity mismatch.")
            return redirect('attendance:issue_book')
        
        issued_count = 0
        for b_id, qty_str in zip(book_ids, quantities):
            try:
                qty = int(qty_str)
            except ValueError:
                qty = 1
            if qty < 1:
                qty = 1
                
            book = get_object_or_404(Book, id=b_id)
            
            if book.available_copies < qty:
                messages.error(
                    request, 
                    f"Not enough copies of '{book.title}'. Available: {book.available_copies}, requested: {qty}."
                )
                return redirect('attendance:issue_book')
            
            for _ in range(qty):
                LibraryRecord.objects.create(
                    student=student_obj,
                    teacher=teacher_obj,
                    book=book,
                    due_date=due_date,
                    remarks=remarks,
                    status='ISSUED'
                )
                book.available_copies -= 1
                book.save()
                issued_count += 1
                
        if issued_count > 0:
            messages.success(request, f"Issued {issued_count} book(s) successfully.")
        return redirect('attendance:manage_library')
        
    # GET request
    students = StudentProfile.objects.all().order_by('name')
    teachers = TeacherProfile.objects.all().order_by('name')
    books = Book.objects.filter(available_copies__gt=0).order_by('title')
    
    return render(request, 'attendance/issue_book.html', {
        'students': students,
        'teachers': teachers,
        'books': books,
        'default_due': (timezone.now().date() + timedelta(days=14)).strftime('%Y-%m-%d'),
    })


@login_required
@transaction.atomic
def return_book(request, record_id):
    if request.user.role not in [User.IS_LIBRARIAN, User.IS_ADMIN]:
        return render(request, 'errors/403.html', status=403)
        
    record = get_object_or_404(LibraryRecord, id=record_id)
    
    if record.status == 'RETURNED':
        messages.info(request, "This record is already marked as returned.")
        return redirect('attendance:manage_library')

    # 1. Update Library Record
    record.status = 'RETURNED'
    record.return_date = timezone.now().date()
    record.save()

    # 2. Increment Available Copies
    book = record.book
    book.available_copies += 1
    book.save()

    # 3. Complete/Reset associated Reserve Request
    if record.student:
        ReserveRequest.objects.filter(
            student=record.student, 
            book=book, 
            status__in=['APPROVED', 'FULFILLED']
        ).update(status='COMPLETED')
    elif record.teacher:
        ReserveRequest.objects.filter(
            teacher=record.teacher, 
            book=book, 
            status__in=['APPROVED', 'FULFILLED']
        ).update(status='COMPLETED')

    borrower_name = record.student.name if record.student else record.teacher.name
    messages.success(request, f"Successfully returned '{book.title}' for {borrower_name}. Reserve status reset!")
    return redirect('attendance:manage_library')

@login_required
def apply_reserve_book(request, book_id):
    if request.user.role not in [User.IS_STUDENT, User.IS_TEACHER]:
        return render(request, 'errors/403.html', status=403)
    
    book = get_object_or_404(Book, id=book_id, is_reserve=True)
    purpose_notes = request.POST.get('purpose_notes', '')

    # Determine profile
    profile = None
    if request.user.role == User.IS_STUDENT:
        try:
            profile = request.user.student_profile
        except StudentProfile.DoesNotExist:
            messages.error(request, "Student profile not found.")
            return redirect('attendance:library_reader_dashboard')
    else:
        try:
            profile = request.user.teacher_profile
        except TeacherProfile.DoesNotExist:
            messages.error(request, "Teacher profile not found.")
            return redirect('attendance:library_reader_dashboard')

    # Build filter kwargs based on role
    user_filter = {'student': profile} if request.user.role == User.IS_STUDENT else {'teacher': profile}

    # 1. Block if there is already a PENDING or APPROVED request for this book
    active_request_exists = ReserveRequest.objects.filter(
        book=book,
        status__in=['PENDING', 'APPROVED'],
        **user_filter
    ).exists()

    if active_request_exists:
        messages.error(request, "You already have an active or pending request for this reserve book.")
        return redirect('attendance:library_reader_dashboard')

    # 2. Block if the user currently holds an active ISSUED record for this book
    currently_issued = LibraryRecord.objects.filter(
        book=book,
        status='ISSUED',
        **user_filter
    ).exists()

    if currently_issued:
        messages.error(request, "You currently have this book checked out. It must be returned first.")
        return redirect('attendance:library_reader_dashboard')

    # 3. Check stock availability
    if book.available_copies < 1:
        messages.error(request, "Sorry, this reserve item is currently out of stock.")
        return redirect('attendance:library_reader_dashboard')

    # Create a fresh application entry (allows re-application after return/completion)
    request_data = {
        'book': book,
        'purpose_notes': purpose_notes,
        'status': 'PENDING'
    }
    request_data.update(user_filter)

    ReserveRequest.objects.create(**request_data)
    messages.success(request, f"Successfully submitted reserve application for '{book.title}'.")
    return redirect('attendance:library_reader_dashboard')

@login_required
@transaction.atomic
def add_book(request):
    """
    Endpoint for the librarian to upload/add books to the system library catalog.
    """
    if request.user.role not in [User.IS_LIBRARIAN, User.IS_ADMIN]:
        return HttpResponseForbidden("Access Blocked.")
        
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        author = request.POST.get('author', '').strip()
        isbn = request.POST.get('isbn', '').strip()
        total_copies_str = request.POST.get('total_copies', '1')
        
        if not title:
            messages.error(request, "Book title is required.")
            return redirect('attendance:manage_library')
            
        try:
            total_copies = int(total_copies_str)
            if total_copies < 1:
                total_copies = 1
        except ValueError:
            total_copies = 1
            
        # Check if book exists (match by unique isbn if available, else exact title)
        book = None
        if isbn:
            book = Book.objects.filter(isbn=isbn).first()
        if not book:
            book = Book.objects.filter(title__iexact=title).first()
            
        if book:
            book.total_copies += total_copies
            book.available_copies += total_copies
            if author:
                book.author = author
            if isbn:
                book.isbn = isbn
            book.save()
            messages.success(request, f"Updated stock for '{book.title}'. Added {total_copies} more copy/copies.")
        else:
            Book.objects.create(
                title=title,
                author=author,
                isbn=isbn if isbn else None,
                total_copies=total_copies,
                available_copies=total_copies
            )
            messages.success(request, f"Successfully uploaded and cataloged '{title}'.")
            
    return redirect('attendance:manage_library')


@login_required
def teacher_dashboard(request):
    """
    Comprehensive Teacher Dashboard processing live operational summaries,
    geospatial session distributions, activity matrices, and dynamic tracking feeds.
    """
    # 1. Structural Role Verification
    if request.user.role != User.IS_TEACHER:
        return HttpResponse("Unauthorized", status=403)
        
    if not hasattr(request.user, 'teacher_profile') or request.user.teacher_profile is None:
        return HttpResponse("Teacher Profile configuration missing.", status=404)
        
    teacher = request.user.teacher_profile
    
    # 2. Base Domain Querysets (Scoped to Active Batch)
    timetable = TimetableEntry.objects.filter(
        batch__is_active=True, teacher=teacher
    ).select_related('course_unit__course__department', 'batch', 'stream')
    
    course_units = CourseUnit.objects.filter(timetableentry__teacher=teacher).distinct()
    
    # --- Multi-Course Assignment Enforcement Guard ---
    has_multiple_units = course_units.count() > 1
    selected_course_unit_id = request.GET.get('selected_course_unit', '')
    
    if has_multiple_units and selected_course_unit_id:
        timetable = timetable.filter(course_unit__code=selected_course_unit_id)
        
    # 3. Extract Filter Parameters from GET for Student Roster & Analytics
    filter_dept = request.GET.get('filter_dept', '')
    filter_course = request.GET.get('filter_course', '')
    filter_stream = request.GET.get('filter_stream', '')
    quick_range = request.GET.get('quick_range', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Select streams assigned directly to this teacher's timetable ruleset
    selectable_streams = Stream.objects.filter(entries__teacher=teacher).distinct()
    
    distinct_courses = set()
    distinct_depts = set()
    for cu in course_units:
        if hasattr(cu, 'course') and cu.course:
            distinct_courses.add(cu.course)
            if hasattr(cu.course, 'department') and cu.course.department:
                distinct_depts.add(cu.course.department)
                
    departments = list(distinct_depts)
    courses = list(distinct_courses)
    
    # 4. Apply Context Criteria Filters on Assigned Student Roster
    students = StudentProfile.objects.filter(course__units__in=course_units).distinct()
    
    if filter_dept:
        students = students.filter(course__department_id=filter_dept)
    if filter_course:
        students = students.filter(course__code=filter_course)
    if filter_stream:
        students = students.filter(stream_id=filter_stream)

    # --- Live Download Data Streaming Channel Interceptor ---
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="student_attendance_roster_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Registration Identification', 'Full Name', 'Department Scope', 'Academic Program Track', 'Class Stream'])
        
        for s in students:
            dept_name = s.course.department.name if s.course.department else "General Academics"
            stream_name = s.stream.name if s.stream else "Unassigned"
            writer.writerow([s.reg_number, s.name, dept_name, s.course.name, stream_name])
            
        return response
        
    total_students_count = students.count()
    active_classes_count = TimetableEntry.objects.filter(batch__is_active=True, teacher=teacher).values('stream').distinct().count()
    
    # Resolve weekday value to match timetable entry day flags
    weekday_map = {0: 'MON', 1: 'TUE', 2: 'WED', 3: 'THU', 4: 'FRI', 5: 'SAT', 6: 'SUN'}
    current_weekday_str = weekday_map[datetime.now().weekday()]
    
    todays_sessions = TimetableEntry.objects.filter(batch__is_active=True, teacher=teacher, day=current_weekday_str)
    todays_sessions_count = todays_sessions.count()
    
    # 5. DYNAMIC METRICS: Filter base records according to UI interactions
    base_records = AttendanceRecord.objects.filter(session__timetable_entry__teacher=teacher)
    
    if has_multiple_units and selected_course_unit_id:
        base_records = base_records.filter(session__timetable_entry__course_unit__code=selected_course_unit_id)
    if filter_dept:
        base_records = base_records.filter(session__timetable_entry__course_unit__course__department_id=filter_dept)
    if filter_course:
        base_records = base_records.filter(session__timetable_entry__course_unit__course__code=filter_course)
    if filter_stream:
        base_records = base_records.filter(session__timetable_entry__stream_id=filter_stream)

    # Process timeframe query parameters matrix
    today = datetime.now().date()
    if quick_range == 'today':
        base_records = base_records.filter(session__date_marked=today)
    elif quick_range == 'week':
        from datetime import timedelta
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        base_records = base_records.filter(session__date_marked__range=[start_of_week, end_of_week])
    elif quick_range == 'month':
        base_records = base_records.filter(session__date_marked__year=today.year, session__date_marked__month=today.month)
    elif start_date and end_date:
        try:
            base_records = base_records.filter(session__date_marked__range=[start_date, end_date])
        except (ValueError, TypeError):
            pass

    # Compute Aggregate Global Attendance Rates on the filtered dataset
    total_records_count = base_records.count()
    present_records_count = base_records.filter(status='PRESENT').count()
    absent_records_count = base_records.filter(status='ABSENT').count()
    
    global_attendance_rate = round((present_records_count / total_records_count) * 100, 1) if total_records_count > 0 else 0
    
    # 6. Generate Dynamic Weekly Attendance Trend Matrix (Respects filters per day)
    weekly_trend_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    weekly_trend_rates = []
    for day_code in ['MON', 'TUE', 'WED', 'THU', 'FRI']:
        day_records = base_records.filter(session__timetable_entry__day=day_code)
        day_total = day_records.count()
        day_present = day_records.filter(status='PRESENT').count()
        day_rate = round((day_present / day_total) * 100, 1) if day_total > 0 else 0
        weekly_trend_rates.append(day_rate)

    # 7. Compile Distribution Analytics Dataset
    distribution_metrics = {
        'present_pct': round((present_records_count / total_records_count) * 100, 1) if total_records_count > 0 else 0,
        'absent_pct': round((absent_records_count / total_records_count) * 100, 1) if total_records_count > 0 else 0,
        'late_pct': 0,
        'custom_range': 0
    }
    
    # 8. Build Recent Timeline Activity Stream
    recent_sessions = AttendanceSession.objects.filter(timetable_entry__teacher=teacher).select_related('timetable_entry__stream').order_by('-id')[:5]
    recent_activity_feed = []
    for session in recent_sessions:
        p_count = session.records.filter(status='PRESENT').count()
        a_count = session.records.filter(status='ABSENT').count()
        recent_activity_feed.append({
            'stream_name': session.timetable_entry.stream.name if session.timetable_entry.stream else "Unknown Session",
            'course_code': session.timetable_entry.course_unit.code,
            'present': p_count,
            'absent': a_count,
            'time': session.date_marked.strftime("%b %d, %Y")
        })

    unit_stats = []
    for cu in course_units:
        p = AttendanceRecord.objects.filter(session__timetable_entry__course_unit=cu, session__timetable_entry__teacher=teacher, status='PRESENT').count()
        t = AttendanceRecord.objects.filter(session__timetable_entry__course_unit=cu, session__timetable_entry__teacher=teacher).count()
        r = round((p / t) * 100, 1) if t > 0 else 0
        unit_stats.append({'name': cu.name, 'rate': r})
        
    unit_names = [u['name'] for u in unit_stats]
    unit_rates = [u['rate'] for u in unit_stats]

    week_start = None
    base_timetable = TimetableEntry.objects.filter(batch__is_active=True, teacher=teacher)
    active_batch = base_timetable.first().batch if base_timetable.exists() else None
    if active_batch:
        week_start = active_batch.week_start_date

    # 9. Identify Timetable Entries Already Marked Today to Prevent Duplicates
    marked_today_ids = list(
        AttendanceSession.objects.filter(
            timetable_entry__teacher=teacher,
            date_marked=today
        ).values_list('timetable_entry_id', flat=True)
    )

    context = {
        'timetable': timetable,
        'students': students,
        'course_units': course_units,
        'departments': departments,
        'courses': courses,
        'selectable_streams': selectable_streams,
        'unit_names': json.dumps(unit_names),
        'unit_rates': json.dumps(unit_rates),
        'week_start': week_start,
        
        # Duplicate Prevention State Tracker
        'marked_today_ids': marked_today_ids,
        'today_date': today,
        
        # Filtering State Retainers
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'quick_range': quick_range,
        'start_date': start_date,
        'end_date': end_date,
        
        # Course Unit Toggle State
        'has_multiple_units': has_multiple_units,
        'selected_course_unit_id': selected_course_unit_id,
        'current_weekday_str': current_weekday_str,
        
        # Summary Counters
        'total_students_count': total_students_count,
        'active_classes_count': active_classes_count,
        'todays_sessions_count': todays_sessions_count,
        'global_attendance_rate': global_attendance_rate,
        'weekly_trend_labels': json.dumps(weekly_trend_labels),
        'weekly_trend_rates': json.dumps(weekly_trend_rates),
        'distribution': distribution_metrics,
        'recent_activity_feed': recent_activity_feed
    }
    return render(request, 'attendance/teacher_dashboard.html', context)
@login_required
def mark_attendance(request, entry_id):
    """
    Renders the student attendance sheet for a valid timetable lecture slot 
    and saves submitted student records while ensuring security controls are met.
    """
    # 1. Base Guard Validation Rules
    if request.user.role != User.IS_TEACHER:
        return HttpResponse("Unauthorized", status=403)
        
    teacher = request.user.teacher_profile
    entry = get_object_or_404(TimetableEntry, id=entry_id, batch__is_active=True)
    
    # Ownership Check: Confirm this specific teacher owns this timetable entry
    if entry.teacher != teacher:
        return HttpResponse("Unauthorized: You are not assigned to this lecture slot.", status=403)
        
    # Day Check: Validate that the teacher is accessing this on the scheduled day
    weekday_map = {0: 'MON', 1: 'TUE', 2: 'WED', 3: 'THU', 4: 'FRI', 5: 'SAT', 6: 'SUN'}
    current_weekday_str = weekday_map[datetime.now().weekday()]
    if entry.day != current_weekday_str:
        messages.error(request, f"Access Locked: This session is scheduled for {entry.get_day_display()}. You can only record attendance on the exact day of execution.")
        return redirect('attendance:teacher_dashboard')
        
    # Duplicate Prevention Check: Notify and block if the lesson has already been submitted today
    today_date = datetime.now().date()
    session_exists = AttendanceSession.objects.filter(timetable_entry=entry, date_marked=today_date).exists()
    if session_exists:
        messages.warning(
            request, 
            f"Notification: Attendance records have already been captured and closed for the "
            f"{entry.course_unit.name} ({entry.stream.name}) lecture slot scheduled today."
        )
        return redirect('attendance:teacher_dashboard')

    # Fetch all students registered under this stream class layout
    students = StudentProfile.objects.filter(stream=entry.stream).order_by('name')

    # 2. Processing POST Submissions
    if request.method == 'POST':
        if AttendanceSession.objects.filter(timetable_entry=entry, date_marked=today_date).exists():
            messages.error(request, "Submission Rejected: This attendance register was already saved by another request.")
            return redirect('attendance:teacher_dashboard')

        lat_raw = request.POST.get('latitude')
        lng_raw = request.POST.get('longitude')
        
        lat = lat_raw.strip() if lat_raw and lat_raw.strip() else None
        lng = lng_raw.strip() if lng_raw and lng_raw.strip() else None
        
        # Instantiate parent Session entry
        session = AttendanceSession.objects.create(
            timetable_entry=entry,
            teacher_latitude=lat,
            teacher_longitude=lng
        )
        
        # Write individual student rows
        for student in students:
            # Captures checkbox value. Checked input outputs 'PRESENT', unselected inputs fallback to 'ABSENT'
            status = request.POST.get(f'student_{student.reg_number}', 'ABSENT')
            
            AttendanceRecord.objects.create(
                session=session,
                student=student,
                status=status
            )
            
        messages.success(request, f"Attendance successfully saved and locked for {entry.course_unit.name} ({entry.stream.name}).")
        return redirect('attendance:teacher_dashboard')

    # 3. Processing GET Requests (Render roster sheet)
    context = {
        'entry': entry,
        'students': students,
        'today_date': today_date
    }
    return render(request, 'attendance/mark_attendance.html', context)

    
import json
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.http import HttpResponse
from .models import User, AttendanceRecord, RoomAllocation, AcademicTerm, LibraryRecord

@login_required
def student_dashboard(request):
    """
    Displays personal attendance summary, historical records, analytics charts,
    hostel allocation, and library borrowing history.
    """
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)
        
    student = request.user.student_profile
    records = AttendanceRecord.objects.filter(student=student).select_related(
        'session__timetable_entry__course_unit__course__department',
        'session__timetable_entry__teacher',
        'session__timetable_entry__stream'
    )
    
    # --- Attendance stats ---
    unit_attendance = {}
    total_present = 0
    total_absent = 0
    
    for rec in records:
        cu_name = rec.session.timetable_entry.course_unit.name
        if cu_name not in unit_attendance:
            unit_attendance[cu_name] = {'present': 0, 'absent': 0}
            
        if rec.status == 'PRESENT':
            unit_attendance[cu_name]['present'] += 1
            total_present += 1
        else:
            unit_attendance[cu_name]['absent'] += 1
            total_absent += 1
    
    unit_names = list(unit_attendance.keys())
    present_counts = [unit_attendance[u]['present'] for u in unit_names]
    absent_counts = [unit_attendance[u]['absent'] for u in unit_names]

    total_sessions = total_present + total_absent
    attendance_percentage = round((total_present / total_sessions) * 100, 1) if total_sessions > 0 else 0.0
    eligible_for_card = attendance_percentage >= 75.0

    # --- Hostel room allocation (current term) ---
    current_term = AcademicTerm.objects.filter(is_current=True).first()
    current_room_allocation = None
    if current_term:
        try:
            current_room_allocation = RoomAllocation.objects.get(student=student, term=current_term)
        except RoomAllocation.DoesNotExist:
            current_room_allocation = None

    # --- Library borrow history (FIXED: use `issue_date` instead of `date_issued`) ---
    library_records = LibraryRecord.objects.filter(student=student).order_by('-issue_date')

    context = {
        'records': records,
        'total_present': total_present,
        'total_absent': total_absent,
        'unit_names': json.dumps(unit_names),
        'present_counts': json.dumps(present_counts),
        'absent_counts': json.dumps(absent_counts),
        'student': student,
        'attendance_percentage': attendance_percentage,
        'eligible_for_card': eligible_for_card,
        'current_room_allocation': current_room_allocation,
        'library_records': library_records,
        'current_term': current_term,
    }
    return render(request, 'attendance/student_dashboard.html', context)


'''
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

# ReportLab core imports for layout and certificate construction
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

@login_required
def download_attendance_card(request):
    """
    Generates and downloads a beautiful PDF Exam Clearance Certificate
    ONLY if the student's attendance matches or exceeds 75%.
    """
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)

    student = request.user.student_profile
    records = AttendanceRecord.objects.filter(student=student)
    
    total_present = records.filter(status='PRESENT').count()
    total_sessions = records.count()
    
    attendance_percentage = (total_present / total_sessions * 100) if total_sessions > 0 else 0
    
    # Strict Security Guard Enforcement
    if attendance_percentage < 75.0:
        return HttpResponse("Forbidden: Ineligible for clearance certificate.", status=403)

    # 1. Setup PDF Document Container in Landscape Layout
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    # 2. Canvas Background Canvas Callback for Visual Certificate Border Styling
    def draw_certificate_frame(canvas, document):
        canvas.saveState()
        
        # Primary Deep Royal Blue Outer Boundary Frame
        canvas.setStrokeColor(colors.HexColor("#1e3a8a"))
        canvas.setLineWidth(5)
        canvas.rect(25, 25, document.pagesize[0] - 50, document.pagesize[1] - 50)
        
        # Secondary Gold/Amber Inner Frame Asset 
        canvas.setStrokeColor(colors.HexColor("#d97706"))
        canvas.setLineWidth(1.5)
        canvas.rect(32, 32, document.pagesize[0] - 64, document.pagesize[1] - 64)
        
        # Abstract Geometric Security Watermark/Seal background placeholder
        canvas.setFillColor(colors.HexColor("#f8fafc"))
        canvas.restoreState()

    # 3. Typography Styles Setup 
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CertTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=26,
        leading=32,
        textColor=colors.HexColor("#1e3a8a"),
        alignment=1  # Centered
    )
    
    subtitle_style = ParagraphStyle(
        'CertSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#b45309"),
        alignment=1,
        spaceAfter=25
    )
    
    body_text_style = ParagraphStyle(
        'CertBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=15,
        leading=24,
        textColor=colors.HexColor("#334155"),
        alignment=1
    )

    # 4. Assembling Content Story Elements
    story = []
    
    # Header Section
    story.append(Spacer(1, 15))
    story.append(Paragraph("UTC BUSHENYI ATTENDANCE HUB", subtitle_style))
    story.append(Paragraph("CERTIFICATE OF EXAMINATION ELIGIBILITY", title_style))
    story.append(Spacer(1, 20))
    
    # Formatted Verification Statement Text Blocks
    dept_name = student.course.department.name if student.course.department else "General Academics"
    statement = (
        f"This is to officially verify and certify that the student listed below has fulfilled "
        f"the mandatory institutional structural attendance requirements framework for the academic session."
    )
    story.append(Paragraph(statement, body_text_style))
    story.append(Spacer(1, 25))
    
    # 5. Core Student Metadata Grid/Table Block Styling
    data_label_style = ParagraphStyle('DataLabel', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor("#1e3a8a"))
    data_val_style = ParagraphStyle('DataVal', fontName='Helvetica', fontSize=12, textColor=colors.HexColor("#1e293b"))
    rate_val_style = ParagraphStyle('RateVal', fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor("#15803d"))

    student_metadata_table_data = [
        [Paragraph("STUDENT NAME:", data_label_style), Paragraph(student.name.upper(), data_val_style),
         Paragraph("REGISTRATION NO:", data_label_style), Paragraph(student.reg_number, data_val_style)],
        [Paragraph("DEPARTMENT:", data_label_style), Paragraph(dept_name, data_val_style),
         Paragraph("PROGRAM TRACK:", data_label_style), Paragraph(f"{student.course.code} - {student.course.name}", data_val_style)],
        [Paragraph("ALLOCATED STREAM:", data_label_style), Paragraph(student.stream.name if student.stream else "Unassigned", data_val_style),
         Paragraph("AGGREGATE ATTENDANCE:", data_label_style), Paragraph(f"{round(attendance_percentage, 1)}% (ELIGIBLE)", rate_val_style)]
    ]
    
    # Render neat custom metrics table layout configuration matrix
    meta_table = Table(student_metadata_table_data, colWidths=[140, 210, 140, 210])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor("#e2e8f0")), # Soft internal dividing dividers
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 45))
    
    # 6. Formal Attestation Signatures Footer Layout block
    sig_line_style = ParagraphStyle('SigLine', fontName='Helvetica', fontSize=10, textColor=colors.HexColor("#64748b"), alignment=1)
    sig_title_style = ParagraphStyle('SigTitle', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#1e3a8a"), alignment=1)

    signatures_layout_matrix = [
        [Paragraph("", sig_line_style), Paragraph("", sig_line_style), Paragraph("", sig_line_style)],
        [Paragraph("<b>___________________________</b>", sig_line_style), 
         Paragraph("<b>[ SYSTEM SEAL ]</b>", sig_title_style), 
         Paragraph("<b>___________________________</b>", sig_line_style)],
        [Paragraph("Academic Registrar Office", sig_title_style), 
         Paragraph("Verified Digitally", sig_line_style), 
         Paragraph("Date of Issuance", sig_title_style)]
    ]
    
    sig_table = Table(signatures_layout_matrix, colWidths=[250, 200, 250])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(sig_table)

    # Build the document passing the decorative background frame
    doc.build(story, onFirstPage=draw_certificate_frame)
    
    # Return processed file stream safely back to the user
    pdf_output = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Exam_Clearance_{student.reg_number}.pdf"'
    response.write(pdf_output)
    return response
'''
import io
import os                                 # <-- added for file existence check
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image                                # <-- added for logo
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

@login_required
def download_attendance_card(request):
    """
    Generates and downloads a beautiful PDF Exam Clearance Certificate
    ONLY if the student's attendance matches or exceeds 75%.
    """
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)

    student = request.user.student_profile
    records = AttendanceRecord.objects.filter(student=student)
    
    total_present = records.filter(status='PRESENT').count()
    total_sessions = records.count()
    
    attendance_percentage = (total_present / total_sessions * 100) if total_sessions > 0 else 0
    
    if attendance_percentage < 75.0:
        return HttpResponse("Forbidden: Ineligible for clearance certificate.", status=403)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    def draw_certificate_frame(canvas, document):
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#1e3a8a"))
        canvas.setLineWidth(5)
        canvas.rect(25, 25, document.pagesize[0] - 50, document.pagesize[1] - 50)
        canvas.setStrokeColor(colors.HexColor("#d97706"))
        canvas.setLineWidth(1.5)
        canvas.rect(32, 32, document.pagesize[0] - 64, document.pagesize[1] - 64)
        canvas.setFillColor(colors.HexColor("#f8fafc"))
        canvas.restoreState()

    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('CertTitle', parent=styles['Normal'],
                                fontName='Helvetica-Bold', fontSize=26, leading=32,
                                textColor=colors.HexColor("#1e3a8a"), alignment=1)
    subtitle_style = ParagraphStyle('CertSubtitle', parent=styles['Normal'],
                                    fontName='Helvetica', fontSize=13, leading=16,
                                    textColor=colors.HexColor("#b45309"), alignment=1, spaceAfter=25)
    body_text_style = ParagraphStyle('CertBody', parent=styles['Normal'],
                                     fontName='Helvetica', fontSize=15, leading=24,
                                     textColor=colors.HexColor("#334155"), alignment=1)

    story = []
    
    # ========== NEW: Top-center logo ==========
    logo_path = 'static/B_logo.png'                   # <-- adjust path if the file is elsewhere
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=140, height=100)   # resize as needed
        logo.hAlign = 'CENTER'
        story.append(logo)
        story.append(Spacer(1, 15))
    # =========================================

    story.append(Spacer(1, 15))
    story.append(Paragraph("UTC BUSHENYI ATTENDANCE HUB", subtitle_style))
    story.append(Paragraph("CERTIFICATE OF EXAMINATION ELIGIBILITY", title_style))
    story.append(Spacer(1, 20))
    
    # (rest of the code remains exactly the same)
    dept_name = student.course.department.name if student.course.department else "General Academics"
    statement = (
        f"This is to officially verify and certify that the student listed below has fulfilled "
        f"the mandatory institutional structural attendance requirements framework for the academic session."
    )
    story.append(Paragraph(statement, body_text_style))
    story.append(Spacer(1, 25))
    
    data_label_style = ParagraphStyle('DataLabel', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor("#1e3a8a"))
    data_val_style = ParagraphStyle('DataVal', fontName='Helvetica', fontSize=12, textColor=colors.HexColor("#1e293b"))
    rate_val_style = ParagraphStyle('RateVal', fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor("#15803d"))

    student_metadata_table_data = [
        [Paragraph("STUDENT NAME:", data_label_style), Paragraph(student.name.upper(), data_val_style),
         Paragraph("REGISTRATION NO:", data_label_style), Paragraph(student.reg_number, data_val_style)],
        [Paragraph("DEPARTMENT:", data_label_style), Paragraph(dept_name, data_val_style),
         Paragraph("PROGRAM TRACK:", data_label_style), Paragraph(f"{student.course.code} - {student.course.name}", data_val_style)],
        [Paragraph("ALLOCATED STREAM:", data_label_style), Paragraph(student.stream.name if student.stream else "Unassigned", data_val_style),
         Paragraph("AGGREGATE ATTENDANCE:", data_label_style), Paragraph(f"{round(attendance_percentage, 1)}% (ELIGIBLE)", rate_val_style)]
    ]
    
    meta_table = Table(student_metadata_table_data, colWidths=[140, 210, 140, 210])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor("#e2e8f0")),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 45))
    
    sig_line_style = ParagraphStyle('SigLine', fontName='Helvetica', fontSize=10, textColor=colors.HexColor("#64748b"), alignment=1)
    sig_title_style = ParagraphStyle('SigTitle', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#1e3a8a"), alignment=1)

    signatures_layout_matrix = [
        [Paragraph("", sig_line_style), Paragraph("", sig_line_style), Paragraph("", sig_line_style)],
        [Paragraph("<b>___________________________</b>", sig_line_style), 
         Paragraph("<b>[ SYSTEM SEAL ]</b>", sig_title_style), 
         Paragraph("<b>___________________________</b>", sig_line_style)],
        [Paragraph("Academic Registrar Office", sig_title_style), 
         Paragraph("Verified Digitally", sig_line_style), 
         Paragraph("Date of Issuance", sig_title_style)]
    ]
    
    sig_table = Table(signatures_layout_matrix, colWidths=[250, 200, 250])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(sig_table)

    doc.build(story, onFirstPage=draw_certificate_frame)
    
    pdf_output = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Exam_Clearance_{student.reg_number}.pdf"'
    response.write(pdf_output)
    return response


@login_required
def download_student_report(request):
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)

    student = request.user.student_profile
    
    # UPDATED: Prefetches course and its pinned department dynamically to minimize lookup loads
    records = AttendanceRecord.objects.filter(student=student).select_related(
        'session__timetable_entry__course_unit__course__department',
        'session__timetable_entry__teacher',
        'session__timetable_entry__stream'
    )

    # Create Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # UPDATED: Appended 'Department' into the reporting headers framework
    headers = ['Department', 'Course Unit', 'Date', 'Status', 'Stream', 'Teacher']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    # Data rows
    for row_num, rec in enumerate(records, start=2):
        session = rec.session
        entry = session.timetable_entry
        
        # Pull the pinned structural department safely
        course_dept = entry.course_unit.course.department
        department_name = course_dept.name if course_dept else "Unassigned"
        
        ws.cell(row=row_num, column=1, value=department_name)
        ws.cell(row=row_num, column=2, value=entry.course_unit.name)
        ws.cell(row=row_num, column=3, value=session.date_marked.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=4, value=rec.status)
        ws.cell(row=row_num, column=5, value=entry.stream.name if entry.stream else "")
        ws.cell(row=row_num, column=6, value=entry.teacher.name)

    # Auto-adjust column widths dynamically
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=\"attendance_report_{student.reg_number}.xlsx\"'
    wb.save(response)
    return response


from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Count, F
from attendance.models import User, RoomAllocation, Hostel, Room, StudentProfile, TimetableEntry


from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Count, F
from attendance.models import User, RoomAllocation, Hostel, Room, StudentProfile, TimetableEntry


from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q, Prefetch
from attendance.models import (
    User, RoomAllocation, Hostel, Room, StudentProfile, 
    TimetableEntry, AcademicTerm, Department, Course, Stream
)

@login_required
def view_lodgings(request):
    user = request.user
    current_term = AcademicTerm.objects.filter(is_current=True).first()
    is_management = user.role in [User.IS_ADMIN, User.IS_WARDEN]
    
    # -------------------------------------------------------------------------
    # HANDLE BULK POST ALLOCATION
    # -------------------------------------------------------------------------
    if request.method == 'POST':
        # SAFE DEFAULTS: Prevents UnboundLocalError under any conditional routing path
        room_id = None
        student_ids = []

        # Strict security guard check
        if not is_management:
            messages.error(request, "Permission Denied: You do not have authority to alter room allocations.")
            return redirect('attendance:view_lodgings')

        # Now safely extract parameters within the validated management block
        room_id = request.POST.get('room_id')
        student_ids = request.POST.getlist('student_ids') 
        
        if not room_id or not student_ids:
            messages.error(request, "Invalid submission. Please select a room and at least one student.")
            return redirect('attendance:view_lodgings')
            
        if not current_term:
            messages.error(request, "Operational Block: No active current academic term set in system settings.")
            return redirect('attendance:view_lodgings')

        try:
            with transaction.atomic():
                # Lock the room row to ensure capacity integrity during bulk assignment
                room = Room.objects.select_for_update().get(id=room_id)
                current_occupancy = RoomAllocation.objects.filter(room=room, term=current_term).count()
                available_slots = room.capacity - current_occupancy
                
                # LINE 881: Now completely safe from UnboundLocalErrors
                if len(student_ids) > available_slots:
                    messages.error(request, f"Allocation Failed: Room has only {available_slots} slots open, but you chose {len(student_ids)} students.")
                    return redirect('attendance:view_lodgings')
                
                allocations_to_create = []
                for s_id in student_ids:
                    student = StudentProfile.objects.get(pk=s_id)
                    
                    # Safe check against the unique_together ('student', 'term') rule
                    if RoomAllocation.objects.filter(student=student, term=current_term).exists():
                        continue
                        
                    allocations_to_create.append(
                        RoomAllocation(
                            student=student,
                            room=room,
                            term=current_term,
                            allocated_by=user
                        )
                    )
                
                if allocations_to_create:
                    RoomAllocation.objects.bulk_create(allocations_to_create)
                    messages.success(request, f"Successfully allocated {len(allocations_to_create)} students.")
                else:
                    messages.warning(request, "No allocations were processed. Selected students may already have rooms this term.")
                    
        except Exception as e:
            messages.error(request, f"Transaction error encountered: {str(e)}")
            
        return redirect('attendance:view_lodgings')
    # -------------------------------------------------------------------------
    # READ DATA QUERIES & PREFETCH SETUP (Remains same for viewing)
    # -------------------------------------------------------------------------
    hostels = Hostel.objects.all()
    
    if current_term:
        allocations_qs = RoomAllocation.objects.filter(term=current_term).select_related('student__stream', 'allocated_by')
        term_filter = Q(allocations__term=current_term)
    else:
        allocations_qs = RoomAllocation.objects.all().select_related('student__stream', 'allocated_by')
        term_filter = Q()
        messages.warning(request, "System Notice: No active academic term is flagged as current. Displaying all historical records.")
    
    if user.role == User.IS_STUDENT:
        allocations_qs = allocations_qs.filter(student__user=user)
    elif user.role == User.IS_TEACHER:
        teacher_streams = TimetableEntry.objects.filter(
            batch__is_active=True, teacher__user=user
        ).values_list('stream_id', flat=True)
        allocations_qs = allocations_qs.filter(student__stream_id__in=teacher_streams)
    elif is_management:
        pass
    else:
        raise PermissionDenied("You do not have access to view student lodgings.")

    departments = Department.objects.all()
    courses = Course.objects.all()
    streams = Stream.objects.all()
    
    rooms_with_occupancy = Room.objects.annotate(
        current_occupancy=Count('allocations', filter=term_filter)
    ).prefetch_related(
        Prefetch('allocations', queryset=allocations_qs, to_attr='term_allocations')
    ).select_related('hostel').order_by('hostel__name', 'name_or_number')

    if is_management:
        allocated_student_ids = RoomAllocation.objects.filter(term=current_term).values_list('student_id', flat=True) if current_term else []
        unallocated_students = StudentProfile.objects.exclude(pk__in=allocated_student_ids).select_related(
            'course__department', 'stream'
        ).order_by('name')
    else:
        unallocated_students = StudentProfile.objects.none()

    context = {
        'hostels': hostels,
        'rooms': rooms_with_occupancy,
        'is_management': is_management,
        'departments': departments,
        'courses': courses,
        'streams': streams,
        'unallocated_students': unallocated_students,
        'current_term': current_term,
    }
    return render(request, 'attendance/view_lodgings.html', context)










@login_required
@transaction.atomic
def allocate_or_reallocate(request):
    # Strict Authorization Guard for Wardens only
    if request.user.role != User.IS_WARDEN:
        return HttpResponse("Unauthorized: Only Wardens can manage room allocations.", status=403)
        
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        room_id = request.POST.get('room_id')
        
        student_profile = get_object_or_404(StudentProfile, reg_number=student_id)
        target_room = get_object_or_404(Room, id=room_id)
        
        # update_or_create handles both fresh assignments and reallocations safely
        RoomAllocation.objects.update_or_create(
            student=student_profile,
            defaults={'room': target_room, 'allocated_by': request.user}
        )
        messages.success(request, f"Successfully allocated room for student {student_profile.name}.")
        return redirect('attendance:view_lodgings')

from django.utils import timezone
from decimal import Decimal

from django.db import models
from django.db.models import Sum, F
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
import calendar

import calendar
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.db import models, transaction
from django.db.models import Sum
from django.utils import timezone
from django.db.models.functions import ExtractMonth, ExtractYear
from attendance.models import (
    User, StudentProfile, AcademicTerm, StudentTermFee, 
    FeePaymentTransaction, StaffPaymentRecord
)

@login_required
def fees_dashboard(request):
    """
    Student fee management page:
    - Management (Admin/Accountant): view all student fee accounts for the current term,
      see transaction history, and record new payments.
    - Student: view only their own fee account and transaction history.
    """
    user = request.user
    current_term = AcademicTerm.objects.filter(is_current=True).first()
    is_management = user.role in [User.IS_ACCOUNTANT, User.IS_ADMIN]

    # ----------------------------------------------------------------------
    # 1. Determine which fee accounts and transactions to display
    # ----------------------------------------------------------------------
    if is_management:
        if current_term:
            fee_accounts = StudentTermFee.objects.filter(term=current_term).select_related('student')
            transactions = FeePaymentTransaction.objects.filter(
                term_fee_account__term=current_term
            ).select_related('term_fee_account__student', 'processed_by')
            all_accounts = StudentTermFee.objects.filter(term=current_term).select_related('student')
        else:
            fee_accounts = StudentTermFee.objects.none()
            transactions = FeePaymentTransaction.objects.none()
            all_accounts = None
            messages.warning(request, "No active academic term is set. Please configure a current term.")
    elif user.role == User.IS_STUDENT:
        # Students see only their own data
        fee_accounts = StudentTermFee.objects.filter(student__user=user).select_related('student', 'term')
        transactions = FeePaymentTransaction.objects.filter(
            term_fee_account__student__user=user
        ).select_related('term_fee_account__student', 'processed_by')
        all_accounts = None
    else:
        raise PermissionDenied("You do not have permission to view student financial records.")

    # ----------------------------------------------------------------------
    # 2. Handle POST – record a new payment (management only)
    # ----------------------------------------------------------------------
    if request.method == 'POST':
        if not is_management:
            raise PermissionDenied("Unauthorized transaction entry blocked.")

        account_id = request.POST.get('account_id')
        amount = request.POST.get('amount')
        ref = request.POST.get('reference_number')
        desc = request.POST.get('description', '')

        try:
            target_account = StudentTermFee.objects.get(id=account_id)
            FeePaymentTransaction.objects.create(
                term_fee_account=target_account,
                amount=amount,
                reference_number=ref,
                description=desc,
                processed_by=user if user.role == User.IS_ACCOUNTANT else None
            )
            messages.success(request, f"Payment successfully recorded for {target_account.student.name}.")
        except Exception as e:
            messages.error(request, f"Failed to log transaction: {str(e)}")

        return redirect('attendance:fees_dashboard')

    # ----------------------------------------------------------------------
    # 3. Context for the template
    # ----------------------------------------------------------------------
    context = {
        'fee_accounts': fee_accounts,
        'transactions': transactions,
        'is_management': is_management,
        'all_accounts': all_accounts,
        'current_term': current_term,
    }
    return render(request, 'attendance/fees_dashboard.html', context)

@login_required
@transaction.atomic
def edit_fee_transaction(request, transaction_id):
    """
    Allows Accountant/Admin to rewrite a fee transaction.
    Commits full current state parameters of the form and modifies corresponding term balances.
    """
    if request.user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        raise PermissionDenied("Unauthorized financial operation.")

    transaction_record = get_object_or_404(FeePaymentTransaction, id=transaction_id)
    account = transaction_record.term_fee_account

    if request.method == 'POST':
        try:
            new_amount = Decimal(request.POST.get('amount', '0.00'))
            new_ref = request.POST.get('reference_number', '').strip()
            new_desc = request.POST.get('description', '')

            # Revert old balance impact if the transaction line was previously confirmed
            if transaction_record.is_confirmed:
                account.total_amount_paid -= transaction_record.amount
                account.total_amount_paid += new_amount
                account.save()

            # Fully overwrite parameters to commit all form fields
            transaction_record.amount = new_amount
            transaction_record.reference_number = new_ref
            transaction_record.description = new_desc
            transaction_record.processed_by = request.user
            transaction_record.save()

            messages.success(request, f"Transaction {new_ref} updated and balances adjusted successfully.")
        except Exception as e:
            messages.error(request, f"Failed to modify ledger row: {str(e)}")
            
    return redirect('attendance:fees_dashboard')


@login_required
@transaction.atomic
def delete_fee_transaction(request, transaction_id):
    """
    Removes a fee transaction permanently and rolls back corresponding paid tracking fields.
    """
    if request.user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        raise PermissionDenied("Unauthorized financial operation.")

    transaction_record = get_object_or_404(FeePaymentTransaction, id=transaction_id)
    account = transaction_record.term_fee_account

    try:
        if transaction_record.is_confirmed:
            account.total_amount_paid -= transaction_record.amount
            account.save()

        ref = transaction_record.reference_number
        transaction_record.delete()
        messages.success(request, f"Transaction reference record {ref} dropped completely.")
    except Exception as e:
        messages.error(request, f"Failed to expunge ledger line: {str(e)}")

    return redirect('attendance:fees_dashboard')


from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect
from django.contrib import messages
from attendance.models import User, AcademicTerm, StaffPaymentRecord

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.shortcuts import render, redirect
from django.contrib import messages
from attendance.models import User, AcademicTerm, StaffPaymentRecord

@login_required
def staff_payments_dashboard(request):
    user = request.user
    current_term = AcademicTerm.objects.filter(is_current=True).first()

    if user.role == User.IS_STUDENT:
        raise PermissionDenied("Students are not permitted to access staff ledger payroll balances.")

    # Rule Verification: Admins/Accountants see everything.
    # Wardens, Teachers see ONLY their own records.
    if user.role in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        payments = StaffPaymentRecord.objects.all().select_related('staff', 'processed_by', 'term').order_by('-payment_date')
        is_management = True
    else:
        payments = StaffPaymentRecord.objects.filter(staff=user).select_related('staff', 'processed_by', 'term').order_by('-payment_date')
        is_management = False

    # --- NEW: Compute stats for the template ---
    total_amount = sum(p.amount for p in payments) if payments else 0
    latest_payment = payments.first() if payments else None
    payment_methods = list(set(p.payment_method for p in payments)) if payments else []

    if request.method == 'POST':
        if user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
            raise PermissionDenied("Unauthorized transaction entry blocked.")
            
        if not current_term:
            messages.error(request, "Cannot process payroll without an active current academic term defined.")
            return redirect('attendance:staff_payments_dashboard')

        staff_id = request.POST.get('staff_id')
        amount = request.POST.get('amount')
        ref = request.POST.get('reference_number')
        desc = request.POST.get('description', '')
        
        try:
            target_staff = User.objects.get(id=staff_id)
            StaffPaymentRecord.objects.create(
                staff=target_staff,
                amount=amount,
                reference_number=ref,
                description=desc,
                term=current_term,
                processed_by=user if user.role == User.IS_ACCOUNTANT else None
            )
            messages.success(request, f"Payment successfully released to {target_staff.username}.")
        except Exception as e:
            messages.error(request, f"Failed to log transaction: {str(e)}")
        return redirect('attendance:staff_payments_dashboard')

    all_staff = User.objects.exclude(role=User.IS_STUDENT) if is_management else None

    context = {
        'payments': payments,
        'is_management': is_management,
        'all_staff': all_staff,
        'current_term': current_term,
        # NEW STATS
        'total_amount': total_amount,
        'latest_payment': latest_payment,
        'payment_methods': payment_methods,
    }
    return render(request, 'attendance/staff_payments_dashboard.html', context)

@login_required
def disburse_payment_view(request):
    """Render the full‑page staff payment disbursement form – only for Accountants/Admins."""
    if request.user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        raise PermissionDenied("Access Denied: Only Accountants and Admins can disburse payments.")
    
    all_staff = User.objects.exclude(role=User.IS_STUDENT).order_by('username')
    return render(request, 'attendance/disburse_payment.html', {'all_staff': all_staff})


@login_required
@transaction.atomic
def edit_staff_payment(request, payment_id):
    """
    Overwrites the complete state parameters for a recorded staff payout.
    """
    if request.user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        raise PermissionDenied("Unauthorized payroll access.")

    payment = get_object_or_404(StaffPaymentRecord, id=payment_id)

    if request.method == 'POST':
        try:
            staff_id = request.POST.get('staff_id')
            amount = request.POST.get('amount')
            ref = request.POST.get('reference_number').strip()
            desc = request.POST.get('description', '')
            term_id = request.POST.get('term_id')

            payment.staff = get_object_or_404(User, id=staff_id)
            payment.amount = Decimal(amount)
            payment.reference_number = ref
            payment.description = desc
            if term_id:
                payment.term = get_object_or_404(AcademicTerm, id=term_id)
            payment.processed_by = request.user
            payment.save()

            messages.success(request, f"Staff disbursement reference {ref} rewritten completely.")
        except Exception as e:
            messages.error(request, f"Failed to modify payload: {str(e)}")

    return redirect('attendance:staff_payments_dashboard')


@login_required
@transaction.atomic
def delete_staff_payment(request, payment_id):
    """
    Deletes a staff payroll payment slip item line.
    """
    if request.user.role not in [User.IS_ACCOUNTANT, User.IS_ADMIN]:
        raise PermissionDenied("Unauthorized payroll access.")

    payment = get_object_or_404(StaffPaymentRecord, id=payment_id)
    try:
        ref = payment.reference_number
        payment.delete()
        messages.success(request, f"Staff payment record {ref} expunged successfully.")
    except Exception as e:
        messages.error(request, f"Failed to delete transaction: {str(e)}")

    return redirect('attendance:staff_payments_dashboard')


@login_required
@transaction.atomic
def record_payment_attempt(request):
    """
    Logs an unconfirmed remittance claim attached to a student's active term account balance.
    """
    if request.user.role != User.IS_ACCOUNTANT:
        return HttpResponse("Unauthorized", status=403)

    current_term = AcademicTerm.objects.filter(is_current=True).first()
    if not current_term:
        messages.error(request, "Operational registration failure: No active academic term defined.")
        return redirect('attendance:fees_dashboard')

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        amount = Decimal(request.POST.get('amount', '0.00'))
        ref_num = request.POST.get('reference_number', '').strip()
        method = request.POST.get('payment_method')

        student_profile = get_object_or_404(StudentProfile, reg_number=student_id)
        fee_account, created = StudentTermFee.objects.get_or_create(
            student=student_profile,
            term=current_term
        )
        
        try:
            FeePaymentTransaction.objects.create(
                term_fee_account=fee_account,
                amount=amount,
                payment_method=method,
                reference_number=ref_num,
                is_confirmed=False
            )
            messages.success(request, f"Payment trace statement logged for {student_profile.name}. Awaiting auditing.")
        except Exception as e:
            messages.error(request, f"Failed to record transaction line: Reference ID might already exist.")

    return redirect('attendance:fees_dashboard')


@login_required
@transaction.atomic
def confirm_student_payment(request, transaction_id):
    """
    Audits an unconfirmed remittance claim and mutates the linked active term balances.
    """
    if request.user.role != User.IS_ACCOUNTANT:
        return HttpResponse("Unauthorized", status=403)

    transaction_record = get_object_or_404(FeePaymentTransaction, id=transaction_id)
    
    if not transaction_record.is_confirmed:
        transaction_record.is_confirmed = True
        transaction_record.date_confirmed = timezone.now()
        transaction_record.processed_by = request.user
        transaction_record.save()
        
        account = transaction_record.term_fee_account
        account.total_amount_paid += transaction_record.amount
        account.save()
        
        messages.success(request, f"Transaction code {transaction_record.reference_number} verified. Balance updated successfully.")
    else:
        messages.warning(request, "This target payment slip has already been settled and audited.")

    return redirect('attendance:fees_dashboard')


from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q
from attendance.models import User, DisciplinaryRecord, StudentProfile, TimetableEntry

@login_required
def disciplinary_dashboard(request):
    user = request.user
    records = DisciplinaryRecord.objects.all().select_related('student__stream', 'reported_by')
    
    # -------------------------------------------------------------------------
    # 1. RECORD VISIBILITY FILTER
    # -------------------------------------------------------------------------
    if user.role in [User.IS_ADMIN, User.IS_WARDEN]:
        # MODIFIED: Warden can now view ALL disciplinary entries without restrictions
        pass
        
    elif user.role == User.IS_TEACHER:
        teacher_streams = TimetableEntry.objects.filter(
            batch__is_active=True, 
            teacher__user=user
        ).values_list('stream_id', flat=True)
        
        records = records.filter(
            Q(student__stream_id__in=teacher_streams) | 
            Q(reported_by=user)
        )
        
    elif user.role == User.IS_STUDENT:
        records = records.filter(student__user=user)
    else:
        records = records.filter(reported_by=user)

    # -------------------------------------------------------------------------
    # 2. ELIGIBLE TARGET STUDENT LIST
    # -------------------------------------------------------------------------
    if user.role in [User.IS_ADMIN, User.IS_WARDEN]:
        # MODIFIED: Warden can view the full student roster to select a target profile
        eligible_students = StudentProfile.objects.all()
    elif user.role == User.IS_TEACHER:
        eligible_students = StudentProfile.objects.all() 
    else:
        eligible_students = StudentProfile.objects.none()

    # -------------------------------------------------------------------------
    # 3. PERMISSION FLAGS SYSTEM
    # -------------------------------------------------------------------------
    # MODIFIED: Added Warden to allowed form-reporting roles (fixes template mismatch)
    can_report = user.role in [User.IS_ADMIN, User.IS_TEACHER, User.IS_WARDEN]
    is_admin = user.role == User.IS_ADMIN

    context = {
        'records': records,
        'students': eligible_students,
        'severities': DisciplinaryRecord.SEVERITY_LEVELS,
        'can_report': can_report,  # Matched variable named inside the HTML template
        'is_admin': is_admin,      # Checked by template to evaluate action rows
    }
    return render(request, 'attendance/disciplinary_dashboard.html', context)



from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from .models import DisciplinaryRecord, StudentProfile, AcademicTerm
from .forms import DisciplinaryEditForm  # we'll create this next

@login_required
def edit_complaint(request, pk):
    record = get_object_or_404(DisciplinaryRecord, pk=pk)
    
    # Permissions: only admin or the original reporter can edit
    if not (request.user.role == User.IS_ADMIN or record.reported_by == request.user):
        messages.error(request, "You are not authorized to edit this complaint.")
        return redirect('attendance:disciplinary_dashboard')
    
    if request.method == 'POST':
        form = DisciplinaryEditForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            messages.success(request, "Complaint record updated successfully.")
            return redirect('attendance:disciplinary_dashboard')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = DisciplinaryEditForm(instance=record)
    
    # Prepare list of students for the dropdown (same as dashboard)
    students = StudentProfile.objects.all().order_by('name')
    terms = AcademicTerm.objects.all().order_by('-academic_year', 'term')
    
    return render(request, 'attendance/disciplinary_edit.html', {
        'form': form,
        'record': record,
        'students': students,
        'terms': terms,
    })

    
@login_required
@transaction.atomic
def add_complaint(request):
    # Security block: Students cannot record complaints
    if request.user.role == User.IS_STUDENT:
        return HttpResponseForbidden("Access Denied: Students cannot log disciplinary complaints.")

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        subject = request.POST.get('subject', '').strip()
        details = request.POST.get('details', '').strip()
        severity = request.POST.get('severity')

        student = get_object_or_404(StudentProfile, reg_number=student_id)
        
        DisciplinaryRecord.objects.create(
            student=student,
            subject=subject,
            details=details,
            severity=severity,
            reported_by=request.user
        )
        messages.success(request, f"Disciplinary report successfully logged against {student.name}.")
    
    return redirect('attendance:disciplinary_dashboard')


@login_required
@transaction.atomic
def delete_complaint(request, record_id):
    record = get_object_or_404(DisciplinaryRecord, id=record_id)
    
    # Ownership Guard: Admin can delete anything, staff can only delete what they reported
    if request.user.role == User.IS_ADMIN or record.reported_by == request.user:
        record.delete()
        messages.success(request, "Disciplinary complaint record successfully expunged.")
        return redirect('attendance:disciplinary_dashboard')
        
    return HttpResponseForbidden("Action Blocked: You can only delete incidents originally logged by yourself.")

    
# ---------- Custom Error Handlers ----------

def custom_page_not_found(request, exception):
    return render(request, 'errors/404.html', status=404)

def custom_permission_denied(request, exception=None):
    return render(request, 'errors/403.html', status=403)

def custom_server_error(request):
    return render(request, 'errors/500.html', status=500)

def custom_bad_request(request, exception=None):
    return render(request, 'errors/400.html', status=400)
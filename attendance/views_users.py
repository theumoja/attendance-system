import json
import csv
from io import BytesIO
from itertools import groupby
from operator import attrgetter

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.urls import reverse

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from attendance.models import (
    User, TimetableEntry, StudentProfile, AttendanceSession,
    AttendanceRecord, CourseUnit
)

@login_required
def home(request):
    """
    Redirects the user to their appropriate dashboard based on their role.
    """
    if request.user.role == User.IS_ADMIN:
        return redirect('attendance:admin_dashboard')
    elif request.user.role == User.IS_TEACHER:
        return redirect('attendance:teacher_dashboard')
    else:
        return redirect('attendance:student_dashboard')


import csv
import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q

# Import all models from models.py
from attendance.models import (
    User, 
    Department, 
    Course, 
    Stream, 
    CourseUnit, 
    TeacherProfile, 
    StudentProfile, 
    TimetableBatch, 
    TimetableEntry, 
    AttendanceSession, 
    AttendanceRecord
)

@login_required
def teacher_dashboard(request):
    """
    Comprehensive Teacher Dashboard processing live operational summaries,
    geospatial session distributions, activity matrices, and dynamic tracking feeds.
    """
    # 1. Structural Role Verification
    if request.user.role != User.IS_TEACHER:
        return HttpResponse("Unauthorized", status=403)
        
    if not hasattr(request.user, 'teacher_profile') or request.user.teacher_profile is None:
        return HttpResponse("Teacher Profile configuration missing.", status=404)
        
    teacher = request.user.teacher_profile
    
    # 2. Base Domain Querysets
    timetable = TimetableEntry.objects.filter(
        batch__is_active=True, teacher=teacher
    ).select_related('course_unit__course__department', 'batch', 'stream')
    
    course_units = CourseUnit.objects.filter(timetableentry__teacher=teacher).distinct()
    
    # --- Multi-Course Assignment Enforcement Guard ---
    has_multiple_units = course_units.count() > 1
    selected_course_unit_id = request.GET.get('selected_course_unit', '')
    
    # If the user has multiple units, narrow down the unlocked entries if a selection is made
    if has_multiple_units and selected_course_unit_id:
        timetable = timetable.filter(course_unit_id=selected_course_unit_id)
        
    # 3. Extract Filter Parameters from GET for Student Roster
    filter_dept = request.GET.get('filter_dept', '')
    filter_course = request.GET.get('filter_course', '')
    filter_stream = request.GET.get('filter_stream', '')
    quick_range = request.GET.get('quick_range', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Using 'entries' backward relation lookup because of related_name='entries' in models.py
    selectable_streams = Stream.objects.filter(entries__teacher=teacher).distinct()
    
    distinct_courses = set()
    distinct_depts = set()
    for cu in course_units:
        if hasattr(cu, 'course') and cu.course:
            distinct_courses.add(cu.course)
            if hasattr(cu.course, 'department') and cu.course.department:
                distinct_depts.add(cu.course.department)
                
    departments = list(distinct_depts)
    courses = list(distinct_courses)
    
    # 4. Apply Context Criteria Filters on Assigned Student Roster
    students = StudentProfile.objects.filter(course__units__in=course_units).distinct()
    
    if filter_dept:
        students = students.filter(course__department_id=filter_dept)
    if filter_course:
        students = students.filter(course__code=filter_course)
    if filter_stream:
        students = students.filter(stream_id=filter_stream)

    # --- Live Download Data Streaming Channel Interceptor ---
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="student_attendance_roster_{datetime.now().strftime("%Y%m%d")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Registration Identification', 'Full Name', 'Department Scope', 'Academic Program Track', 'Class Stream'])
        
        for s in students:
            dept_name = s.course.department.name if s.course.department else "General Academics"
            stream_name = s.stream.name if s.stream else "Unassigned"
            writer.writerow([s.reg_number, s.name, dept_name, s.course.name, stream_name])
            
        return response
        
    total_students_count = students.count()
    active_classes_count = TimetableEntry.objects.filter(batch__is_active=True, teacher=teacher).values('stream').distinct().count()
    
    # Resolve weekday value to look up systemic locks
    weekday_map = {0: 'MON', 1: 'TUE', 2: 'WED', 3: 'THU', 4: 'FRI', 5: 'SAT', 6: 'SUN'}
    current_weekday_str = weekday_map[datetime.now().weekday()]
    
    todays_sessions = TimetableEntry.objects.filter(batch__is_active=True, teacher=teacher, day=current_weekday_str)
    todays_sessions_count = todays_sessions.count()
    
    # 5. Compute Aggregate Global Attendance Rates
    total_records = AttendanceRecord.objects.filter(session__timetable_entry__teacher=teacher)
    total_records_count = total_records.count()
    present_records_count = total_records.filter(status='PRESENT').count()
    absent_records_count = total_records.filter(status='ABSENT').count()
    
    global_attendance_rate = round((present_records_count / total_records_count) * 100, 1) if total_records_count > 0 else 0
    
    # 6. Generate Weekly Attendance Trend Matrix
    weekly_trend_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    weekly_trend_rates = []
    for day_code in ['MON', 'TUE', 'WED', 'THU', 'FRI']:
        day_total = AttendanceRecord.objects.filter(session__timetable_entry__teacher=teacher, session__timetable_entry__day=day_code).count()
        day_present = AttendanceRecord.objects.filter(session__timetable_entry__teacher=teacher, session__timetable_entry__day=day_code, status='PRESENT').count()
        day_rate = round((day_present / day_total) * 100, 1) if day_total > 0 else 0
        weekly_trend_rates.append(day_rate)

    # 7. Compile Distribution Analytics Dataset
    distribution_metrics = {
        'present_pct': round((present_records_count / total_records_count) * 100, 1) if total_records_count > 0 else 0,
        'absent_pct': round((absent_records_count / total_records_count) * 100, 1) if total_records_count > 0 else 0,
        'late_pct': 0,
        'custom_range': 0
    }
    
    # 8. Build Recent Timeline Activity Stream
    recent_sessions = AttendanceSession.objects.filter(timetable_entry__teacher=teacher).select_related('timetable_entry__stream').order_by('-id')[:5]
    recent_activity_feed = []
    for session in recent_sessions:
        p_count = session.records.filter(status='PRESENT').count()
        a_count = session.records.filter(status='ABSENT').count()
        time_str = datetime.now().strftime("%H:%M")
        recent_activity_feed.append({
            'stream_name': session.timetable_entry.stream.name if session.timetable_entry.stream else "Unknown Session",
            'present': p_count,
            'absent': a_count,
            'time': time_str
        })

    unit_stats = []
    for cu in course_units:
        p = AttendanceRecord.objects.filter(session__timetable_entry__course_unit=cu, session__timetable_entry__teacher=teacher, status='PRESENT').count()
        t = AttendanceRecord.objects.filter(session__timetable_entry__course_unit=cu, session__timetable_entry__teacher=teacher).count()
        r = round((p / t) * 100, 1) if t > 0 else 0
        unit_stats.append({'name': cu.name, 'rate': r})
        
    unit_names = [u['name'] for u in unit_stats]
    unit_rates = [u['rate'] for u in unit_stats]

    week_start = None
    base_timetable = TimetableEntry.objects.filter(batch__is_active=True, teacher=teacher)
    active_batch = base_timetable.first().batch if base_timetable.exists() else None
    if active_batch:
        week_start = active_batch.week_start_date

    # 9. Identify Timetable Entries Already Marked Today to Prevent Duplicates
    today_date = datetime.now().date()
    marked_today_ids = list(
        AttendanceSession.objects.filter(
            timetable_entry__teacher=teacher,
            date_marked=today_date
        ).values_list('timetable_entry_id', flat=True)
    )

    context = {
        'timetable': timetable,
        'students': students,
        'course_units': course_units,
        'departments': departments,
        'courses': courses,
        'selectable_streams': selectable_streams,
        'unit_names': json.dumps(unit_names),
        'unit_rates': json.dumps(unit_rates),
        'week_start': week_start,
        
        # Duplicate Prevention State Tracker
        'marked_today_ids': marked_today_ids,
        
        # Filtering State Retainers
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'quick_range': quick_range,
        'start_date': start_date,
        'end_date': end_date,
        
        # Course Unit Toggle State
        'has_multiple_units': has_multiple_units,
        'selected_course_unit_id': selected_course_unit_id,
        'current_weekday_str': current_weekday_str,
        
        # Summary Counters
        'total_students_count': total_students_count,
        'active_classes_count': active_classes_count,
        'todays_sessions_count': todays_sessions_count,
        'global_attendance_rate': global_attendance_rate,
        'weekly_trend_labels': json.dumps(weekly_trend_labels),
        'weekly_trend_rates': json.dumps(weekly_trend_rates),
        'distribution': distribution_metrics,
        'recent_activity_feed': recent_activity_feed
    }
    return render(request, 'attendance/teacher_dashboard.html', context)


from datetime import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse
from .models import User, TimetableEntry, AttendanceSession, StudentProfile, AttendanceRecord

@login_required
def mark_attendance(request, entry_id):
    """
    Renders the student attendance sheet for a valid timetable lecture slot 
    and saves submitted student records while ensuring security controls are met.
    """
    # 1. Base Guard Validation Rules
    if request.user.role != User.IS_TEACHER:
        return HttpResponse("Unauthorized", status=403)
        
    teacher = request.user.teacher_profile
    entry = get_object_or_404(TimetableEntry, id=entry_id, batch__is_active=True)
    
    # Ownership Check: Confirm this specific teacher owns this timetable entry
    if entry.teacher != teacher:
        return HttpResponse("Unauthorized: You are not assigned to this lecture slot.", status=403)
        
    # Day Check: Validate that the teacher is accessing this on the scheduled day
    weekday_map = {0: 'MON', 1: 'TUE', 2: 'WED', 3: 'THU', 4: 'FRI', 5: 'SAT', 6: 'SUN'}
    current_weekday_str = weekday_map[datetime.now().weekday()]
    if entry.day != current_weekday_str:
        messages.error(request, f"Access Locked: This session is scheduled for {entry.get_day_display()}. You can only record attendance on the exact day of execution.")
        return redirect('attendance:teacher_dashboard')
        
    # Duplicate Prevention Check: Notify and block if the lesson has already been submitted today
    today_date = datetime.now().date()
    session_exists = AttendanceSession.objects.filter(timetable_entry=entry, date_marked=today_date).exists()
    if session_exists:
        messages.warning(
            request, 
            f"Notification: Attendance records have already been captured and closed for the "
            f"{entry.course_unit.name} ({entry.stream.name}) lecture slot scheduled today."
        )
        return redirect('attendance:teacher_dashboard')

    # Fetch all students registered under this stream class layout
    students = StudentProfile.objects.filter(stream=entry.stream).order_by('name')

    # 2. Processing POST Submissions (Now capturing native form payloads perfectly)
    if request.method == 'POST':
        # Double-check right before database writes to prevent race conditions or form resubmissions
        if AttendanceSession.objects.filter(timetable_entry=entry, date_marked=today_date).exists():
            messages.error(request, "Submission Rejected: This attendance register was already saved by another request.")
            return redirect('attendance:teacher_dashboard')

        # Read and sanitize coordinates safely from the native form submission [source: 2]
        lat_raw = request.POST.get('latitude')
        lng_raw = request.POST.get('longitude')
        
        # Ensure empty form strings switch cleanly to None to avoid DB validation type crashes
        lat = lat_raw.strip() if lat_raw and lat_raw.strip() else None
        lng = lng_raw.strip() if lng_raw and lng_raw.strip() else None
        
        # Instantiate parent Session entry
        session = AttendanceSession.objects.create(
            timetable_entry=entry,
            teacher_latitude=lat,
            teacher_longitude=lng
        )
        
        # Write individual student rows
        for student in students:
            # Captures standard field parameters format: name="student_{{ s.reg_number }}" [source: 2]
            # Default fallback state is set to 'ABSENT' if unselected or missing
            status = request.POST.get(f'student_{student.reg_number}', 'ABSENT')
            
            AttendanceRecord.objects.create(
                session=session,
                student=student,
                status=status
            )
            
        messages.success(request, f"Attendance successfully saved and locked for {entry.course_unit.name} ({entry.stream.name}).")
        return redirect('attendance:teacher_dashboard')

    # 3. Processing GET Requests (Render standard list layout)
    context = {
        'entry': entry,
        'students': students,
        'today_date': today_date
    }
    return render(request, 'attendance/mark_attendance.html', context)

@login_required
def student_dashboard(request):
    """
    Displays personal attendance summary, historical records, analytics charts,
    and handles attendance card clearance thresholds.
    """
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)
        
    student = request.user.student_profile
    records = AttendanceRecord.objects.filter(student=student).select_related('session__timetable_entry__course_unit')
    
    unit_attendance = {}
    total_present = 0
    total_absent = 0
    
    for rec in records:
        cu_name = rec.session.timetable_entry.course_unit.name
        if cu_name not in unit_attendance:
            unit_attendance[cu_name] = {'present': 0, 'absent': 0}
            
        if rec.status == 'PRESENT':
            unit_attendance[cu_name]['present'] += 1
            total_present += 1
        else:
            unit_attendance[cu_name]['absent'] += 1
            total_absent += 1
    
    unit_names = list(unit_attendance.keys())
    present_counts = [unit_attendance[u]['present'] for u in unit_names]
    absent_counts = [unit_attendance[u]['absent'] for u in unit_names]

    # --- NEW: Compute overall attendance percentage and threshold clearance ---
    total_sessions = total_present + total_absent
    attendance_percentage = round((total_present / total_sessions) * 100, 1) if total_sessions > 0 else 0.0
    eligible_for_card = attendance_percentage >= 75.0

    context = {
        'records': records,
        'total_present': total_present,
        'total_absent': total_absent,
        'unit_names': json.dumps(unit_names),
        'present_counts': json.dumps(present_counts),
        'absent_counts': json.dumps(absent_counts),
        'student': student,
        'attendance_percentage': attendance_percentage,  # Passed to UI
        'eligible_for_card': eligible_for_card,          # Passed to UI
    }
    return render(request, 'attendance/student_dashboard.html', context)


import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

# ReportLab core imports for layout and certificate construction
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

@login_required
def download_attendance_card(request):
    """
    Generates and downloads a beautiful PDF Exam Clearance Certificate
    ONLY if the student's attendance matches or exceeds 75%.
    """
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)

    student = request.user.student_profile
    records = AttendanceRecord.objects.filter(student=student)
    
    total_present = records.filter(status='PRESENT').count()
    total_sessions = records.count()
    
    attendance_percentage = (total_present / total_sessions * 100) if total_sessions > 0 else 0
    
    # Strict Security Guard Enforcement
    if attendance_percentage < 75.0:
        return HttpResponse("Forbidden: Ineligible for clearance certificate.", status=403)

    # 1. Setup PDF Document Container in Landscape Layout
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=40,
        rightMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    # 2. Canvas Background Canvas Callback for Visual Certificate Border Styling
    def draw_certificate_frame(canvas, document):
        canvas.saveState()
        
        # Primary Deep Royal Blue Outer Boundary Frame
        canvas.setStrokeColor(colors.HexColor("#1e3a8a"))
        canvas.setLineWidth(5)
        canvas.rect(25, 25, document.pagesize[0] - 50, document.pagesize[1] - 50)
        
        # Secondary Gold/Amber Inner Frame Asset 
        canvas.setStrokeColor(colors.HexColor("#d97706"))
        canvas.setLineWidth(1.5)
        canvas.rect(32, 32, document.pagesize[0] - 64, document.pagesize[1] - 64)
        
        # Abstract Geometric Security Watermark/Seal background placeholder
        canvas.setFillColor(colors.HexColor("#f8fafc"))
        canvas.restoreState()

    # 3. Typography Styles Setup 
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CertTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=26,
        leading=32,
        textColor=colors.HexColor("#1e3a8a"),
        alignment=1  # Centered
    )
    
    subtitle_style = ParagraphStyle(
        'CertSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor("#b45309"),
        alignment=1,
        spaceAfter=25
    )
    
    body_text_style = ParagraphStyle(
        'CertBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=15,
        leading=24,
        textColor=colors.HexColor("#334155"),
        alignment=1
    )

    # 4. Assembling Content Story Elements
    story = []
    
    # Header Section
    story.append(Spacer(1, 15))
    story.append(Paragraph("UTC BUSHENYI ATTENDANCE HUB", subtitle_style))
    story.append(Paragraph("CERTIFICATE OF EXAMINATION ELIGIBILITY", title_style))
    story.append(Spacer(1, 20))
    
    # Formatted Verification Statement Text Blocks
    dept_name = student.course.department.name if student.course.department else "General Academics"
    statement = (
        f"This is to officially verify and certify that the student listed below has fulfilled "
        f"the mandatory institutional structural attendance requirements framework for the academic session."
    )
    story.append(Paragraph(statement, body_text_style))
    story.append(Spacer(1, 25))
    
    # 5. Core Student Metadata Grid/Table Block Styling
    data_label_style = ParagraphStyle('DataLabel', fontName='Helvetica-Bold', fontSize=12, textColor=colors.HexColor("#1e3a8a"))
    data_val_style = ParagraphStyle('DataVal', fontName='Helvetica', fontSize=12, textColor=colors.HexColor("#1e293b"))
    rate_val_style = ParagraphStyle('RateVal', fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor("#15803d"))

    student_metadata_table_data = [
        [Paragraph("STUDENT NAME:", data_label_style), Paragraph(student.name.upper(), data_val_style),
         Paragraph("REGISTRATION NO:", data_label_style), Paragraph(student.reg_number, data_val_style)],
        [Paragraph("DEPARTMENT:", data_label_style), Paragraph(dept_name, data_val_style),
         Paragraph("PROGRAM TRACK:", data_label_style), Paragraph(f"{student.course.code} - {student.course.name}", data_val_style)],
        [Paragraph("ALLOCATED STREAM:", data_label_style), Paragraph(student.stream.name if student.stream else "Unassigned", data_val_style),
         Paragraph("AGGREGATE ATTENDANCE:", data_label_style), Paragraph(f"{round(attendance_percentage, 1)}% (ELIGIBLE)", rate_val_style)]
    ]
    
    # Render neat custom metrics table layout configuration matrix
    meta_table = Table(student_metadata_table_data, colWidths=[140, 210, 140, 210])
    meta_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('LINEBELOW', (0,0), (-1,-2), 0.5, colors.HexColor("#e2e8f0")), # Soft internal dividing dividers
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 45))
    
    # 6. Formal Attestation Signatures Footer Layout block
    sig_line_style = ParagraphStyle('SigLine', fontName='Helvetica', fontSize=10, textColor=colors.HexColor("#64748b"), alignment=1)
    sig_title_style = ParagraphStyle('SigTitle', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#1e3a8a"), alignment=1)

    signatures_layout_matrix = [
        [Paragraph("", sig_line_style), Paragraph("", sig_line_style), Paragraph("", sig_line_style)],
        [Paragraph("<b>___________________________</b>", sig_line_style), 
         Paragraph("<b>[ SYSTEM SEAL ]</b>", sig_title_style), 
         Paragraph("<b>___________________________</b>", sig_line_style)],
        [Paragraph("Academic Registrar Office", sig_title_style), 
         Paragraph("Verified Digitally", sig_line_style), 
         Paragraph("Date of Issuance", sig_title_style)]
    ]
    
    sig_table = Table(signatures_layout_matrix, colWidths=[250, 200, 250])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(sig_table)

    # Build the document passing the decorative background frame
    doc.build(story, onFirstPage=draw_certificate_frame)
    
    # Return processed file stream safely back to the user
    pdf_output = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Exam_Clearance_{student.reg_number}.pdf"'
    response.write(pdf_output)
    return response




@login_required
def download_student_report(request):
    if request.user.role != User.IS_STUDENT:
        return HttpResponse("Unauthorized", status=403)

    student = request.user.student_profile
    
    # UPDATED: Prefetches course and its pinned department dynamically to minimize lookup loads
    records = AttendanceRecord.objects.filter(student=student).select_related(
        'session__timetable_entry__course_unit__course__department',
        'session__timetable_entry__teacher',
        'session__timetable_entry__stream'
    )

    # Create Excel workbook using openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # UPDATED: Appended 'Department' into the reporting headers framework
    headers = ['Department', 'Course Unit', 'Date', 'Status', 'Stream', 'Teacher']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    # Data rows
    for row_num, rec in enumerate(records, start=2):
        session = rec.session
        entry = session.timetable_entry
        
        # Pull the pinned structural department safely
        course_dept = entry.course_unit.course.department
        department_name = course_dept.name if course_dept else "Unassigned"
        
        ws.cell(row=row_num, column=1, value=department_name)
        ws.cell(row=row_num, column=2, value=entry.course_unit.name)
        ws.cell(row=row_num, column=3, value=session.date_marked.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=4, value=rec.status)
        ws.cell(row=row_num, column=5, value=entry.stream.name if entry.stream else "")
        ws.cell(row=row_num, column=6, value=entry.teacher.name)

    # Auto-adjust column widths dynamically
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=\"attendance_report_{student.reg_number}.xlsx\"'
    wb.save(response)
    return response

# ---------- Custom Error Handlers ----------

def custom_page_not_found(request, exception):
    return render(request, 'errors/404.html', status=404)

def custom_permission_denied(request, exception=None):
    return render(request, 'errors/403.html', status=403)

def custom_server_error(request):
    return render(request, 'errors/500.html', status=500)

def custom_bad_request(request, exception=None):
    return render(request, 'errors/400.html', status=400)
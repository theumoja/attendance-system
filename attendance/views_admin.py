import csv
import io
import json
import secrets
from datetime import datetime, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Count, Q
from attendance.models import *


def generate_secure_password():
    return secrets.token_urlsafe(8)

@login_required
@transaction.atomic
def manage_streams(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        name = request.POST.get('stream_name', '').strip()
        course_code = request.POST.get('course_code', '').strip()
        
        if name and course_code:
            try:
                course = Course.objects.get(code=course_code)
                Stream.objects.create(name=name, course=course)
                messages.success(request, f"Stream '{name}' successfully registered.")
            except Course.DoesNotExist:
                messages.error(request, "The specified parent course layout does not exist.")
            except Exception as e:
                messages.error(request, f"Error building structural layout: {str(e)}")
        else:
            messages.error(request, "All required input layout items must be filled.")
        return redirect('attendance:manage_streams')

    streams = Stream.objects.select_related('course').all()
    courses = Course.objects.all()
    return render(request, 'attendance/manage_streams.html', {
        'streams': streams,
        'courses': courses
    })

@login_required
@transaction.atomic
def edit_stream(request, stream_id):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    stream_obj = get_object_or_404(Stream, id=stream_id)
    if request.method == 'POST':
        name = request.POST.get('stream_name', '').strip()
        course_code = request.POST.get('course_code', '').strip()
        
        if name and course_code:
            try:
                course = Course.objects.get(code=course_code)
                stream_obj.name = name
                stream_obj.course = course
                stream_obj.save()
                messages.success(request, "Structural stream modifications compiled securely.")
                return redirect('attendance:manage_streams')
            except Course.DoesNotExist:
                messages.error(request, "Target parent course layout context does not exist.")
        else:
            messages.error(request, "Fields cannot be blank.")

    courses = Course.objects.all()
    return render(request, 'attendance/edit_stream.html', {
        'stream': stream_obj,
        'courses': courses
    })

@login_required
@transaction.atomic
def delete_stream(request, stream_id):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    stream_obj = get_object_or_404(Stream, id=stream_id)
    name = stream_obj.name
    stream_obj.delete()
    messages.success(request, f"Stream framework structure '{name}' detached successfully.")
    return redirect('attendance:manage_streams')

@login_required
@transaction.atomic
def bulk_upload_streams(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        if not csv_file.name.endswith('.csv'):
            messages.error(request, "Data error: Expected standard comma-delimited flat CSV file structure.")
            return redirect('attendance:bulk_upload_streams')

        try:
            stream_data = csv_file.read().decode('utf-8')
            io_string = io.StringIO(stream_data)
            reader = csv.reader(io_string)
            next(reader, None)

            success_count = 0
            for row in reader:
                if not row or len(row) < 2:
                    continue
                stream_name = row[0].strip()
                course_code = row[1].strip()

                if stream_name and course_code:
                    course = Course.objects.filter(code=course_code).first()
                    if course:
                        Stream.objects.get_or_create(name=stream_name, course=course)
                        success_count += 1

            messages.success(request, f"Ingested {success_count} unique structural academic stream mappings.")
            return redirect('attendance:manage_streams')
        except Exception as e:
            messages.error(request, f"Parser exception detected: {str(e)}")

    return render(request, 'attendance/bulk_upload_streams.html')

# =========================================================================
# 1. TEACHERS MANAGEMENT
# =========================================================================

@login_required
@transaction.atomic
def manage_teachers(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'single' or ('name' in request.POST and 'email' in request.POST):
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
            course_codes = request.POST.getlist('courses')
            
            if name and email:
                username = email.split('@')[0]
                password = generate_secure_password()
                user, created = User.objects.get_or_create(email=email, defaults={
                    'username': username,
                    'role': User.IS_TEACHER,
                    'raw_password_archive': password
                })
                if created:
                    user.set_password(password)
                    user.save()
                    teacher = TeacherProfile.objects.create(user=user, name=name)
                    if course_codes:
                        teacher.courses.set(Course.objects.filter(code__in=course_codes))
                    messages.success(request, f"Teacher account created for {name}.")
                else:
                    messages.warning(request, "A user with this email already exists.")
            return redirect('attendance:manage_teachers')

        elif action == 'bulk' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
            next(reader, None)

            for row in reader:
                if len(row) >= 2:
                    name, email = row[0].strip(), row[1].strip()
                    username = email.split('@')[0]
                    password = generate_secure_password()
                    user, created = User.objects.get_or_create(email=email, defaults={
                        'username': username,
                        'role': User.IS_TEACHER,
                        'raw_password_archive': password
                    })
                    if created:
                        user.set_password(password)
                        user.save()
                        TeacherProfile.objects.create(user=user, name=name)
            return redirect('attendance:export_credentials', role_type='teachers')

    teachers = TeacherProfile.objects.select_related('user').prefetch_related('courses').all()
    all_courses = Course.objects.all()
    return render(request, 'attendance/manage_teachers.html', {
        'teachers': teachers, 
        'all_courses': all_courses,
        'active_courses_list': all_courses
    })


@login_required
@transaction.atomic
def edit_teacher(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    if request.method == 'POST':
        teacher.name = request.POST.get('name', '').strip()
        teacher.user.email = request.POST.get('email', '').strip()
        teacher.user.save()
        teacher.save()
        
        course_codes = request.POST.getlist('courses')
        teacher.courses.set(Course.objects.filter(code__in=course_codes))
        messages.success(request, "Teacher record updated successfully.")
        return redirect('attendance:manage_teachers')
    
    all_courses = Course.objects.all()
    return render(request, 'attendance/edit_teacher.html', {'teacher': teacher, 'all_courses': all_courses})


@login_required
@transaction.atomic
def delete_teacher(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    teacher = get_object_or_404(TeacherProfile, pk=pk)
    user = teacher.user
    teacher.delete()
    user.delete()
    messages.success(request, "Teacher record permanently deleted.")
    return redirect('attendance:manage_teachers')


# =========================================================================
# 2. STUDENTS MANAGEMENT
# =========================================================================

@login_required
@transaction.atomic
def manage_students(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'single':
            name = (request.POST.get('students_name') or request.POST.get('name', '')).strip()
            reg_number = (request.POST.get('registration_number') or request.POST.get('reg_number', '')).strip()
            email = request.POST.get('email', '').strip()
            course_code = request.POST.get('course', '').strip() 
            stream_id = request.POST.get('stream', '').strip()
            
            if name and reg_number and email and course_code:
                try:
                    course = Course.objects.get(code=course_code)
                    stream = Stream.objects.get(id=stream_id) if stream_id else None
                    password = generate_secure_password()
                    username = email.split('@')[0]
                    
                    user = User.objects.create_user(
                        username=username, 
                        email=email, 
                        password=password,
                        role=User.IS_STUDENT
                    )
                    user.raw_password_archive = password
                    user.save()
                    
                    StudentProfile.objects.create(
                        reg_number=reg_number,
                        user=user,
                        name=name,
                        course=course,
                        stream=stream
                    )
                    messages.success(request, f"Student {name} successfully registered.")
                except Course.DoesNotExist:
                    messages.error(request, "Selected course does not exist.")
                except Stream.DoesNotExist:
                    messages.error(request, "Selected stream does not exist.")
                except Exception as e:
                    messages.error(request, f"Error occurred: {str(e)}")
            
            return redirect('attendance:manage_students')

    students = StudentProfile.objects.select_related('user', 'course', 'stream').all()
    courses = Course.objects.all()
    streams = Stream.objects.all()
    
    return render(request, 'attendance/manage_students.html', {
        'students': students,
        'active_courses_list': courses,
        'active_streams_list': streams
    })

@login_required
@transaction.atomic
def edit_student(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    student = get_object_or_404(StudentProfile, pk=pk)
    if request.method == 'POST':
        student.name = (request.POST.get('students_name') or request.POST.get('name', '')).strip()
        student.reg_number = (request.POST.get('registration_number') or request.POST.get('reg_number', '')).strip()
        student.user.email = request.POST.get('email', '').strip()
        
        course_code = request.POST.get('course', '').strip()
        stream_id = request.POST.get('stream', '').strip()
        try:
            student.course = Course.objects.get(code=course_code)
        except Course.DoesNotExist:
            pass

        try:
            student.stream = Stream.objects.get(id=stream_id) if stream_id else None
        except Stream.DoesNotExist:
            pass
            
        student.user.save()
        student.save()
        messages.success(request, "Student records updated cleanly.")
        return redirect('attendance:manage_students')
        
    courses = Course.objects.all()
    streams = Stream.objects.filter(course=student.course)
    return render(request, 'attendance/edit_student.html', {
        'student': student, 
        'active_courses_list': courses,
        'streams': streams
    })


@login_required
@transaction.atomic
def delete_student(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    student = get_object_or_404(StudentProfile, pk=pk)
    user = student.user
    student.delete()
    user.delete()
    messages.success(request, "Student deleted from database repository.")
    return redirect('attendance:manage_students')


# =========================================================================
# NEW. DEPARTMENTS MANAGEMENT
# =========================================================================

@login_required
@transaction.atomic
def manage_departments(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        name = request.POST.get('department_name', '').strip()
        if name:
            Department.objects.get_or_create(name=name)
            messages.success(request, f"Department '{name}' successfully integrated.")
        else:
            messages.error(request, "Department name value cannot be blank.")
        return redirect('attendance:manage_departments')

    departments = Department.objects.all()
    return render(request, 'attendance/manage_departments.html', {'departments': departments})

@login_required
def add_department(request):
    """
    Handles the structural ingestion and creation of a new academic department.
    """
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        dept_name = request.POST.get('department_name', '').strip()
        
        if not dept_name:
            messages.error(request, "Department name cannot be empty.")
            return redirect('attendance:manage_departments')
            
        # Check for duplication framework arrays
        if Department.objects.filter(name__iexact=dept_name).exists():
            messages.error(request, f"The department '{dept_name}' already exists.")
            return redirect('attendance:manage_departments')
            
        # Create structural track entry
        Department.objects.create(name=dept_name)
        messages.success(request, f"Department '{dept_name}' successfully configured.")
        
    return redirect('attendance:manage_departments')

    
@login_required
@transaction.atomic
def edit_department(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('department_name', '').strip()
        if name:
            department.name = name
            department.save()
            messages.success(request, "Department structural identifier modified successfully.")
            return redirect('attendance:manage_departments')
        else:
            messages.error(request, "Department fields cannot be blank.")
            
    return render(request, 'attendance/edit_department.html', {'department': department})


@login_required
@transaction.atomic
def delete_department(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    department = get_object_or_404(Department, pk=pk)
    name = department.name
    department.delete()
    messages.success(request, f"Department '{name}' cleanly detached from application context.")
    return redirect('attendance:manage_departments')


# =========================================================================
# 3. COURSES MANAGEMENT (UPDATED FOR DEPARTMENTS)
# =========================================================================

@login_required
@transaction.atomic
def manage_courses(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'single' or 'course_code' in request.POST:
            code = (request.POST.get('course_code') or request.POST.get('code', '')).strip()
            name = (request.POST.get('course_name') or request.POST.get('name', '')).strip()
            department_id = request.POST.get('department', '').strip()
            
            if code and name:
                dept = Department.objects.filter(id=department_id).first() if department_id else None
                course, created = Course.objects.get_or_create(code=code, defaults={'name': name, 'department': dept})
                if not created:
                    course.name = name
                    course.department = dept
                    course.save()
                messages.success(request, f"Course Program '{code}' integrated.")
            return redirect('attendance:manage_courses')

        elif action == 'bulk' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
            next(reader, None)

            for row in reader:
                if len(row) >= 2:
                    c_code, c_name = row[0].strip(), row[1].strip()
                    if c_code:  # Safety guard for bulk uploads
                        Course.objects.get_or_create(code=c_code, defaults={'name': c_name})
            return redirect('attendance:manage_courses')

    # FIX: Exclude any corrupted empty or null text rows from the stream
    courses = Course.objects.select_related('department').exclude(code="").exclude(code__isnull=True)
    departments = Department.objects.all()
    
    return render(request, 'attendance/manage_courses.html', {
        'courses': courses,
        'departments': departments
    })

@login_required
@transaction.atomic
def edit_course(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.code = (request.POST.get('course_code') or request.POST.get('code', '')).strip()
        course.name = (request.POST.get('course_name') or request.POST.get('name', '')).strip()
        department_id = request.POST.get('department', '').strip()
        
        if department_id:
            course.department = Department.objects.filter(id=department_id).first()
        else:
            course.department = None
            
        course.save()
        messages.success(request, "Course program updated successfully.")
        return redirect('attendance:manage_courses')
        
    departments = Department.objects.all()
    return render(request, 'attendance/edit_course.html', {
        'course': course,
        'departments': departments
    })


@login_required
@transaction.atomic
def delete_course(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
    
    course = get_object_or_404(Course, pk=pk)
    course.delete()
    messages.success(request, "Course program completely wiped from scope variables.")
    return redirect('attendance:manage_courses')


# =========================================================================
# 4. COURSE UNITS MANAGEMENT
# =========================================================================
@login_required
@transaction.atomic
def manage_course_units(request):
    print(f"\n--- [DEBUG] Enter manage_course_units View ---")
    print(f"[DEBUG] Request Method: {request.method} | User: {request.user.username} | Role: {getattr(request.user, 'role', 'None')}")

    # 1. Authorization Guard
    if request.user.role != User.IS_ADMIN:
        print(f"[DEBUG] Access Denied: User role {request.user.role} is not IS_ADMIN.")
        return HttpResponse("Unauthorized", status=403)

    # 2. Handle Form Submissions
    if request.method == 'POST':
        action = request.POST.get('action')
        print(f"[DEBUG] POST Action received: '{action}'")
        
        # --- SINGLE COURSE UNIT CREATION / UPDATE ---
        if action == 'single' or 'course_unit_code' in request.POST:
            print(f"[DEBUG] Routing to SINGLE course unit processing")
            code = (request.POST.get('course_unit_code') or request.POST.get('code', '')).strip()
            name = (request.POST.get('course_unit_name') or request.POST.get('name', '')).strip()
            course_code = request.POST.get('course_code', '').strip()
            
            print(f"[DEBUG] Parsed Single Data -> Code: '{code}', Name: '{name}', Parent Course Code: '{course_code}'")

            if code and name and course_code:
                try:
                    course = Course.objects.get(code=course_code)
                    print(f"[DEBUG] Found parent Course instance PK '{course.pk}' for code '{course_code}'")
                    
                    obj, created = CourseUnit.objects.update_or_create(
                        code=code, 
                        defaults={'name': name, 'course': course}
                    )
                    
                    if created:
                        print(f"[DEBUG] SUCCESS: Created NEW CourseUnit PK '{obj.pk}' (Code: {code})")
                        messages.success(request, f"Unit module '{code}' created and linked directly to parent.")
                    else:
                        print(f"[DEBUG] SUCCESS: UPDATED existing CourseUnit PK '{obj.pk}' (Code: {code})")
                        messages.success(request, f"Unit module '{code}' successfully updated.")
                        
                except Course.DoesNotExist:
                    print(f"[DEBUG] ERROR: Parent Course with code '{course_code}' does not exist in database.")
                    messages.error(request, "Failed addition: Specified Parent Course doesn't exist.")
            else:
                print(f"[DEBUG] ERROR: Validation failed. One or more required fields were evaluated as empty strings.")
                messages.error(request, "Failed addition: Missing required fields.")
                
            return redirect('attendance:manage_course_units')

        # --- BULK CSV UPLOAD ---
        elif action == 'bulk' and request.FILES.get('csv_file'):
            csv_file = request.FILES['csv_file']
            print(f"[DEBUG] Routing to BULK CSV processing. File name: '{csv_file.name}'")
            
            decoded_file = csv_file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
            next(reader, None)  # Skip CSV header row

            saved_count = 0
            failed_rows = []

            for index, row in enumerate(reader, start=2):
                if len(row) >= 4:
                    c_code = row[0].strip()
                    cu_code = row[2].strip()
                    cu_name = row[3].strip()
                    
                    if not c_code or not cu_code or not cu_name:
                        failed_rows.append(f"Row {index} (Missing fields)")
                        continue

                    try:
                        course = Course.objects.get(code=c_code)
                        CourseUnit.objects.update_or_create(
                            code=cu_code, 
                            defaults={'name': cu_name, 'course': course}
                        )
                        saved_count += 1
                    except Course.DoesNotExist:
                        failed_rows.append(f"Row {index} (Course code '{c_code}' not found)")
                        continue
                else:
                    failed_rows.append(f"Row {index} (Malformed column structure)")

            if saved_count > 0:
                messages.success(request, f"Successfully processed {saved_count} course unit(s).")
            if failed_rows:
                messages.error(request, f"Skipped rows due to errors: {', '.join(failed_rows)}")

            return redirect('attendance:manage_course_units')

    # 3. Handle GET Request / Fetch View Context
    course_units = CourseUnit.objects.select_related('course').all()
    courses = Course.objects.all()
    
    for c in courses:
        c.course_code = c.code
        c.course_name = c.name

    return render(request, 'attendance/manage_course_units.html', {
        'course_units': course_units, 
        'courses': courses,
        'active_courses_list': courses
    })



@login_required
@transaction.atomic
def edit_course_unit(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    unit = get_object_or_404(CourseUnit, pk=pk)
    if request.method == 'POST':
        unit.code = (request.POST.get('course_unit_code') or request.POST.get('code', '')).strip()
        unit.name = (request.POST.get('course_unit_name') or request.POST.get('name', '')).strip()
        
        course_code = request.POST.get('course_code', '').strip()
        try:
            unit.course = Course.objects.get(code=course_code)
        except Course.DoesNotExist:
            pass
            
        unit.save()
        messages.success(request, "Subject Module context definitions modified successfully.")
        return redirect('attendance:manage_course_units')
        
    courses = Course.objects.all()
    return render(request, 'attendance/edit_course_unit.html', {'unit': unit, 'active_courses_list': courses})


@login_required
@transaction.atomic
def delete_course_unit(request, pk):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    unit = get_object_or_404(CourseUnit, pk=pk)
    unit.delete()
    messages.success(request, "Module index row cleanly detached and purged.")
    return redirect('attendance:manage_course_units')


# =========================================================================
# 5. CORE ADMINISTRATIVE DASHBOARD & UTILITIES
# =========================================================================

@login_required
def admin_dashboard(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    total_courses = Course.objects.count()
    total_teachers = TeacherProfile.objects.count()
    total_students = StudentProfile.objects.count()
    total_present = AttendanceRecord.objects.filter(status='PRESENT').count()
    total_records = AttendanceRecord.objects.count()
    overall_rate = round((total_present / total_records) * 100, 2) if total_records > 0 else 0

    teacher_stats = []
    teachers = TeacherProfile.objects.all()
    for t in teachers:
        present = AttendanceRecord.objects.filter(
            session__timetable_entry__teacher=t, status='PRESENT'
        ).count()
        total = AttendanceRecord.objects.filter(session__timetable_entry__teacher=t).count()
        rate = round((present / total) * 100, 2) if total > 0 else 0
        teacher_stats.append({'name': t.name, 'rate': rate})
    
    teacher_names = [ts['name'] for ts in teacher_stats]
    teacher_rates = [ts['rate'] for ts in teacher_stats]

    student_stats = StudentProfile.objects.annotate(
        present=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        total=Count('attendancerecord')
    ).order_by('-total')[:10]
    
    student_labels = [s.name for s in student_stats]
    student_present = [s.present for s in student_stats]
    student_absent = [s.total - s.present for s in student_stats]

    stream_counts = Stream.objects.select_related('course').annotate(
        student_count=Count('students')
    ).order_by('name')

    context = {
        'total_courses': total_courses,
        'total_teachers': total_teachers,
        'total_students': total_students,
        'overall_rate': overall_rate,
        'teacher_names': json.dumps(teacher_names),
        'teacher_rates': json.dumps(teacher_rates),
        'student_labels': json.dumps(student_labels),
        'student_present': json.dumps(student_present),
        'student_absent': json.dumps(student_absent),
        'stream_counts': stream_counts,
    }
    return render(request, 'attendance/admin_dashboard.html', context)


@login_required
@transaction.atomic
def bulk_upload_courses(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
        next(reader, None)

        for row in reader:
            if len(row) >= 4:
                c_code, c_name, cu_code, cu_name = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip()
                course, _ = Course.objects.get_or_create(code=c_code, defaults={'name': c_name})
                CourseUnit.objects.get_or_create(code=cu_code, defaults={'name': cu_name, 'course': course})
        return redirect('attendance:admin_dashboard')
    return render(request, 'attendance/upload_courses.html')


@login_required
@transaction.atomic
def bulk_upload_teachers(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
        next(reader, None)

        for row in reader:
            if len(row) >= 2:
                name, email = row[0].strip(), row[1].strip()
                username = email.split('@')[0]
                password = generate_secure_password()

                user, created = User.objects.get_or_create(email=email, defaults={
                    'username': username,
                    'role': User.IS_TEACHER,
                    'raw_password_archive': password
                })
                if created:
                    user.set_password(password)
                    user.save()
                    TeacherProfile.objects.create(user=user, name=name)
        return redirect('attendance:export_credentials', role_type='teachers')
    return render(request, 'attendance/upload_teachers.html')


@login_required
@transaction.atomic
def bulk_upload_students(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
        decoded_file = csv_file.read().decode('utf-8')
        reader = csv.reader(io.StringIO(decoded_file), delimiter=',')
        next(reader, None)

        for row in reader:
            if len(row) >= 5:
                name, reg_num, email, course_code, stream_name = row[0].strip(), row[1].strip(), row[2].strip(), row[3].strip(), row[4].strip()
                try:
                    course = Course.objects.get(code=course_code)
                    
                    from attendance.models import Stream
                    stream_obj, _ = Stream.objects.get_or_create(name=stream_name, course=course)
                    
                    password = generate_secure_password()
                    user, created = User.objects.get_or_create(email=email, defaults={
                        'username': reg_num.replace('/', '_'),
                        'role': User.IS_STUDENT,
                        'raw_password_archive': password
                    })
                    if created:
                        user.set_password(password)
                        user.save()
                        StudentProfile.objects.create(user=user, reg_number=reg_num, name=name, course=course, stream=stream_obj)
                except Course.DoesNotExist:
                    continue
        return redirect('attendance:export_credentials', role_type='students')
    return render(request, 'attendance/upload_students.html')

import json
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from datetime import datetime
from .models import TimetableBatch, TimetableEntry, CourseUnit, TeacherProfile, Stream
import json
import json
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.contrib import messages
from datetime import datetime
from .models import TimetableBatch, TimetableEntry, CourseUnit, TeacherProfile, Stream

import json
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from attendance.models import User, Stream, TimetableBatch, TimetableEntry, CourseUnit, TeacherProfile


@login_required
def manage_timetable(request):
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # Fetch active batch
    batch = TimetableBatch.objects.filter(is_active=True, is_revoked=False).order_by('-uploaded_at').first()
    
    # Handle optional clear action to wipe old test data and start fresh
    if request.GET.get('action') == 'clear_all' and batch:
        TimetableEntry.objects.filter(batch=batch).delete()
        messages.success(request, "All old timetable entries have been cleared. You now have a clean slate.")
        return redirect('attendance:manage_timetable')

    streams = Stream.objects.select_related('course').all()
    
    # Strictly filter and count entries that actually belong to each stream for this batch
    streams_list = []
    for stream in streams:
        entry_count = 0
        if batch:
            entry_count = TimetableEntry.objects.filter(batch=batch, stream=stream).count()
            
        streams_list.append({
            'id': stream.id,
            'name': stream.name,
            'course_name': stream.course.name if stream.course else "No Course Assigned",
            'has_timetable': entry_count > 0,  # True only if entries exist specifically for this class
            'entry_count': entry_count
        })

    return render(request, 'attendance/manage_timetable_list.html', {
        'streams_list': streams_list,
        'batch': batch
    })

@login_required
@transaction.atomic
def upload_timetable(request, stream_id):
    """
    Interactive Timetable Matrix Editor focused strictly on an individual stream/class.
    Omits stream selection from cells and saves all configurations under the current stream.
    """
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    stream_obj = get_object_or_404(Stream, id=stream_id)

    # Retrieve or initialize active batch configuration
    batch = TimetableBatch.objects.filter(is_active=True, is_revoked=False).order_by('-uploaded_at').first()
    if not batch:
        batch = TimetableBatch.objects.create(week_start_date=datetime.now().date(), is_active=True)

    DAYS = [code for code, _ in TimetableEntry.DAYS_OF_WEEK]

    if request.method == 'POST':
        # Wipe existing records for this stream under the current batch to overwrite cleanly
        TimetableEntry.objects.filter(batch=batch, stream=stream_obj).delete()

        row_indexes = request.POST.getlist('row_index')
        
        for idx in row_indexes:
            start_time = request.POST.get(f'start_time_{idx}')
            end_time = request.POST.get(f'end_time_{idx}')
            
            if not start_time or not end_time:
                continue

            for day_code in DAYS:
                cu_code = request.POST.get(f'cu_{idx}_{day_code}')
                teacher_id = request.POST.get(f'teacher_{idx}_{day_code}')

                # Save record if a course unit is chosen for this slot
                if cu_code:
                    try:
                        course_unit = CourseUnit.objects.get(code=cu_code)
                        teacher = TeacherProfile.objects.get(id=teacher_id) if teacher_id else None
                        
                        if teacher:
                            TimetableEntry.objects.create(
                                batch=batch,
                                day=day_code,
                                start_time=start_time,
                                end_time=end_time,
                                course_unit=course_unit,
                                teacher=teacher,
                                stream=stream_obj # Contextually assigned class stream
                            )
                    except Exception as e:
                        pass

        messages.success(request, f"Timetable configuration saved successfully for class '{stream_obj.name}'.")
        return redirect('attendance:upload_timetable', stream_id=stream_obj.id)

    # GET Process: Construct matrix rows specifically for this stream
    existing_entries = TimetableEntry.objects.filter(batch=batch, stream=stream_obj).select_related('course_unit', 'teacher')
    
    # Group entries by unique time slots
    time_slots_map = {}
    for entry in existing_entries:
        key = (entry.start_time.strftime('%H:%M'), entry.end_time.strftime('%H:%M'))
        if key not in time_slots_map:
            time_slots_map[key] = {}
        time_slots_map[key][entry.day] = entry

    matrix_rows = []
    counter = 0
    for (start, end), day_entries in time_slots_map.items():
        slots = []
        for day_code in DAYS:
            slots.append({
                'day_code': day_code,
                'entry': day_entries.get(day_code)
            })
        matrix_rows.append({
            'index': counter,
            'start': start,
            'end': end,
            'slots': slots
        })
        counter += 1

    # Generate qualified instructor mapping arrays
    cu_lecturer_map = {}
    for cu in CourseUnit.objects.all():
        teachers = TeacherProfile.objects.filter(courses=cu.course)
        if not teachers.exists():
            teachers = TeacherProfile.objects.all()
        cu_lecturer_map[str(cu.code)] = [
            {'id': t.id, 'name': t.name} for t in teachers
        ]

    context = {
        'stream': stream_obj,
        'matrix_rows': matrix_rows,
        'days': TimetableEntry.DAYS_OF_WEEK,
        'course_units': CourseUnit.objects.all(),
        'cu_lecturer_map_json': json.dumps(cu_lecturer_map),
    }
    return render(request, 'attendance/upload_timetable.html', context)

def download_template(request, template_type):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{template_type}_template.csv"'

    if template_type == 'courses':
        content = "course_code,course_name,course_unit_code,course_unit_name\nC001,Computer Science,CS101,Programming Basics\nC001,Computer Science,CS102,Data Structures"
    elif template_type == 'teachers':
        content = "teachers_name,email\nJohn Doe,john@example.com"
    elif template_type == 'students':
        content = "students_name,registration_number,email,course,stream_name\nAlice Smith,2024/001,alice@example.com,CIT,Year 1 CIT A"
    elif template_type == 'timetable':
        content = "day,start_time,end_time,course_unit_code,teacher_email,stream_name\nMON,08:30,10:00,CIT101,james.muwonge@utc.ac.ug,Year 1 CIT A"
    else:
        content = ""

    response.write(content)
    return response


# =========================================================================
# 6. CREDENTIAL EXPORTS ENGINE (EXCEL OUTPUT FORMAT)
# =========================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

@login_required
def export_credentials(request, role_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Credentials"

    headers = ['Identifier/Email', 'Name', 'Generated Password']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='d3d3d3', end_color='d3d3d3', fill_type='solid')

    row_num = 2
    if role_type == 'teachers':
        profiles = TeacherProfile.objects.select_related('user').all()
        for p in profiles:
            ws.cell(row=row_num, column=1, value=p.user.email)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.user.raw_password_archive)
            row_num += 1
    elif role_type == 'students':
        profiles = StudentProfile.objects.select_related('user').all()
        for p in profiles:
            ws.cell(row=row_num, column=1, value=p.reg_number)
            ws.cell(row=row_num, column=2, value=p.name)
            ws.cell(row=row_num, column=3, value=p.user.raw_password_archive)
            row_num += 1

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        column = col[0].column_letter
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{role_type}_credentials.xlsx"'
    wb.save(response)
    return response


# =========================================================================
# NEW. ADMINISTRATIVE REPORTS VIEW
# =========================================================================


@login_required
def admin_report_page(request):
    # Enforce Admin-only access
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    departments = Department.objects.all()
    
    # Extract structural query constraint parameter directly from backend request arrays
    selected_dept_id = request.GET.get('filter_dept')
    
    # Start with base layout mapping matching all active structural course matrices
    courses = Course.objects.select_related('department').all()
    
    # Execute structural filtering directly inside database processing layers if constraint exists
    if selected_dept_id:
        courses = courses.filter(department_id=selected_dept_id)
        try:
            selected_dept_id = int(selected_dept_id)  # Standardize type matching for UI template rules
        except ValueError:
            selected_dept_id = None

    return render(request, 'attendance/admin_report.html', {
        'departments': departments,
        'courses': courses,
        'selected_dept_id': selected_dept_id
    })

'''
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
import json
from attendance.models import User, Stream, AttendanceRecord, AttendanceSession, StudentProfile

@login_required
def analytics_dashboard(request):

    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # Extract optional filter parameters from template request scopes
    selected_stream_id = request.GET.get('stream')
    
    # Base query sets targeting core transactional datasets
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related(
        'timetable_entry__teacher__user', 
        'timetable_entry__stream'
    ).order_by('-date_marked')
    
    # Dynamic application of scoping filters if requested by the end-user
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 1. Compute Live Summary Statistics Metrics Matrix
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
        'late_rate': 0  # Maintained at 0 since model STATUS_CHOICES only specifies PRESENT/ABSENT
    }

    # 2. Compile Lecturer Location Audit Log Dataset Dynamically
    audit_logs = []
    # Fetching the top 20 latest session instances to prevent UI buffer layout breaks
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  # Placeholder asset as IP capture hooks are not defined inside current models schema
            "verified": has_coords  # Flag verified true strictly if geolocation values exist
        })

    # 3. Pull Top Students Metrics Arrays via Database Annotations
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT'))
    ).filter(total_rec__gt=0)
    
    top_students_raw = []
    for s in students:
        rate = (s.present_rec / s.total_rec) * 100
        top_students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "rate_num": rate,
            "rate": f"{round(rate, 1)}%"
        })
    
    # Sort and slice dataset to grab top performing arrays
    top_students = sorted(top_students_raw, key=lambda x: x['rate_num'], reverse=True)[:5]

    # 4. Generate Aggregated Payload Engines for ChartJS
    # Chart A: Global Present vs Absent distribution count vectors
    chart_dist_data = [present_count, absent_count]
    
    # Chart B: Generate stream breakdown labels and cross-cutting attendance percentage matrices
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    # Context compilation pass payload arrays safely out into frontend DOM nodes
    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_students': top_students,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. Summary Statistics Metrics
    stats = {
        'total_records': 10,
        'present_rate': 90,
        'absent_rate': 10,
        'late_rate': 0
    }

    # 2. Lecturer Location Audit Log Dataset
    audit_logs = [
        {"date": "2026-06-26 @ 15:21", "class": "NDEE Weekend", "lecturer": "laban.omenyo.ai@gmail.com", "location": "0.3375, 32.5943 (±193642m)", "ip": "41.210.154.15", "verified": True},
        {"date": "2026-06-25 @ 21:45", "class": "NDME A", "lecturer": "technologiestescal@gmail.com", "location": "-0.6195, 30.6697 (±65m)", "ip": "41.210.154.139", "verified": True},
        {"date": "2026-05-18 @ 16:20", "class": "NDEE Weekend", "lecturer": "technologiestescal@gmail.com", "location": "-0.5470, 30.2457 (±130m)", "ip": "154.72.206.194", "verified": True},
        {"date": "2026-05-10 @ 09:13", "class": "NDEE Weekend", "lecturer": "technologiestescal@gmail.com", "location": "-0.5469, 30.2457 (±68m)", "ip": "41.210.146.63", "verified": True},
        {"date": "2026-05-10 @ 22:06", "class": "NDME A", "lecturer": "technologiestescal@gmail.com", "location": "Not captured", "ip": "41.210.146.63", "verified": True},
        {"date": "2026-05-10 @ 21:54", "class": "NDME A", "lecturer": "technologiestescal@gmail.com", "location": "-0.5502, 30.2441 (±30m)", "ip": "41.210.146.63", "verified": False},
        {"date": "2026-05-10 @ 12:36", "class": "NDA A", "lecturer": "technologiestescal@gmail.com", "location": "-0.5469, 30.2457 (±61m)", "ip": "154.72.206.194", "verified": True},
        {"date": "2026-05-10 @ 12:35", "class": "NDA A", "lecturer": "technologiestescal@gmail.com", "location": "Not captured", "ip": "154.72.206.194", "verified": False},
        {"date": "2026-05-08 @ 22:54", "class": "NDEE Weekend", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
        {"date": "2026-05-09 @ 22:33", "class": "NDA A", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
        {"date": "2026-05-09 @ 10:57", "class": "NDEE Weekend", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
        {"date": "2026-05-09 @ 10:39", "class": "NDME A", "lecturer": "—", "location": "Not captured", "ip": "—", "verified": False},
    ]

    # 3. Top Students Placeholder Data Structure
    top_students = [
        {"name": "Mukasa John", "reg": "2025/NDEE/021", "class": "NDEE Weekend", "rate": "98%"},
        {"name": "Ainomugisha Dianah", "reg": "2025/NDME/104", "class": "NDME A", "rate": "96%"},
        {"name": "Mwesigye Brian", "reg": "2025/NDA/008", "class": "NDA A", "rate": "95%"},
    ]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_students': top_students,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)



'''
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from datetime import datetime, timedelta
import json
import csv
from attendance.models import User, Department, Course, Stream, AttendanceRecord, AttendanceSession, StudentProfile

import csv
import json
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.core.exceptions import FieldError

# Assuming your models are imported like this:
# from .models import User, Department, Course, Stream, StudentProfile, AttendanceRecord, AttendanceSession

@login_required
def analytics_dashboard(request):
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 6. INTERCEPT GLOBAL CSV EXPORT REQUEST
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            writer.writerow([record.student.name, record.student.reg_number, stream_name, record.status, date_val])
        return response

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. ADDED/UPDATED: Lecturer Attendance Submission Compliance Report Scope Filtering
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    
    compliance_sessions = AttendanceSession.objects.all()
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    # Apply structural Department, Course, and Stream scope constraints dynamically
    if filter_stream:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=filter_stream)
        timetable_entries = timetable_entries.filter(stream_id=filter_stream)
    elif filter_course:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__code=filter_course)
        timetable_entries = timetable_entries.filter(stream__course__code=filter_course)
    elif filter_dept:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__department_id=filter_dept)
        timetable_entries = timetable_entries.filter(stream__course__department_id=filter_dept)

    if selected_stream_id and selected_stream_id != 'all':
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=selected_stream_id)
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    total_system_sessions = compliance_sessions.count()
    teacher_counts = compliance_sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Data Processing
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # Detailed reporting fields context
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str and start_date_str != 'None':
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str and end_date_str != 'None':
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 6. INTERCEPT GLOBAL CSV EXPORT REQUEST
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            writer.writerow([record.student.name, record.student.reg_number, stream_name, record.status, date_val])
        return response

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. RESILIENT CASCADING LOGIC FOR LECTURER COMPLIANCE REPORT
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    
    compliance_sessions = AttendanceSession.objects.all()
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    # Apply dynamic scoping constraints based on user dropdown selections
    if filter_stream:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=filter_stream)
        timetable_entries = timetable_entries.filter(stream_id=filter_stream)
    elif filter_course:
        try:
            # Look up via standard Stream -> Course relation path
            compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__code=filter_course)
            timetable_entries = timetable_entries.filter(stream__course__code=filter_course)
        except Exception:
            try:
                # Look up via direct Course schema field link (if any)
                compliance_sessions = compliance_sessions.filter(timetable_entry__course__code=filter_course)
                timetable_entries = timetable_entries.filter(course__code=filter_course)
            except Exception:
                pass
    elif filter_dept:
        try:
            # Tier 1 Lookup: Query via Stream -> Course -> Department link mapping
            compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__department_id=filter_dept)
            timetable_entries = timetable_entries.filter(stream__course__department_id=filter_dept)
            
            if not timetable_entries.exists():
                raise FieldError()
        except (FieldError, Exception):
            try:
                # Tier 2 Lookup: Query via direct Course relationship on Timetable model
                compliance_sessions = AttendanceSession.objects.filter(timetable_entry__course__department_id=filter_dept)
                timetable_entries = TimetableEntry.objects.select_related('teacher__user').filter(course__department_id=filter_dept)
                
                if not timetable_entries.exists():
                    raise FieldError()
            except (FieldError, Exception):
                try:
                    # Tier 3 Lookup: Filter by the target Lecturer's home Department structural assignments
                    compliance_sessions = AttendanceSession.objects.filter(
                        Q(timetable_entry__teacher__department_id=filter_dept) | 
                        Q(timetable_entry__teacher__user__department_id=filter_dept)
                    )
                    timetable_entries = TimetableEntry.objects.select_related('teacher__user').filter(
                        Q(teacher__department_id=filter_dept) | 
                        Q(teacher__user__department_id=filter_dept)
                    )
                except (FieldError, Exception):
                    # Tier 4 Safe-Recovery Fallback: Maintain active structural roster list context
                    compliance_sessions = AttendanceSession.objects.all()
                    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()

    # Apply global secondary dashboard overrides if set
    if selected_stream_id and selected_stream_id != 'all':
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=selected_stream_id)
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    # Compute explicit metrics (ensures teachers show up even with 0 submission values)
    total_system_sessions = compliance_sessions.count()
    teacher_counts = compliance_sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Visual Processing Data Arrays
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # Structural cascades UI framework mapping parameters
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 6. INTERCEPT GLOBAL CSV EXPORT REQUEST
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            writer.writerow([record.student.name, record.student.reg_number, stream_name, record.status, date_val])
        return response

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. ADDED/UPDATED: Lecturer Attendance Submission Compliance Report Scope Filtering
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    
    compliance_sessions = AttendanceSession.objects.all()
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    # Apply structural Department, Course, and Stream scope constraints dynamically
    if filter_stream:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=filter_stream)
        timetable_entries = timetable_entries.filter(stream_id=filter_stream)
    elif filter_course:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__code=filter_course)
        timetable_entries = timetable_entries.filter(stream__course__code=filter_course)
    elif filter_dept:
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream__course__department_id=filter_dept)
        timetable_entries = timetable_entries.filter(stream__course__department_id=filter_dept)

    if selected_stream_id and selected_stream_id != 'all':
        compliance_sessions = compliance_sessions.filter(timetable_entry__stream_id=selected_stream_id)
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    total_system_sessions = compliance_sessions.count()
    teacher_counts = compliance_sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Data Processing
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # Detailed reporting fields context
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # 1. CAPTURE DETAILED REPORT FILTERS
    filter_dept = request.GET.get('filter_dept')
    filter_course = request.GET.get('filter_course')
    filter_stream = request.GET.get('filter_stream')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    quick_range = request.GET.get('quick_range')

    # Base entity dropdown datasets supporting cascading display lists
    all_departments = Department.objects.all()
    selectable_courses = Course.objects.all()
    selectable_streams = Stream.objects.all()

    if filter_dept:
        selectable_courses = selectable_courses.filter(department_id=filter_dept)
    if filter_course:
        selectable_streams = selectable_streams.filter(course__code=filter_course)

    # 2. RESOLVE TIME DATA WINDOW CONSTRAINTS
    today = datetime.now().date()
    start_date = None
    end_date = None

    if quick_range == 'today':
        start_date = today
        end_date = today
    elif quick_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif quick_range == 'month':
        start_date = today - timedelta(days=30)
        end_date = today
    else:
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Build cross-cutting performance filter arrays
    record_date_filter = Q()
    if start_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__gte=start_date)
    if end_date:
        record_date_filter &= Q(attendancerecord__session__date_marked__lte=end_date)

    # 3. COMPILE COMPLETE DETAILED STUDENT POPULATION ROSTER
    report_students_queryset = StudentProfile.objects.select_related('stream', 'course__department').all()
    
    if filter_stream:
        report_students_queryset = report_students_queryset.filter(stream_id=filter_stream)
    elif filter_course:
        report_students_queryset = report_students_queryset.filter(course__code=filter_course)
    elif filter_dept:
        report_students_queryset = report_students_queryset.filter(course__department_id=filter_dept)

    # Calculate absolute percentages over selected date parameters
    annotated_students = report_students_queryset.annotate(
        total_filtered_records=Count('attendancerecord', filter=record_date_filter),
        present_filtered_records=Count('attendancerecord', filter=record_date_filter & Q(attendancerecord__status='PRESENT'))
    ).order_by('name')

    detailed_student_reports = []
    for s in annotated_students:
        tot = s.total_filtered_records
        pres = s.present_filtered_records
        percentage_val = round((pres / tot) * 100, 1) if tot > 0 else 0.0
        
        detailed_student_reports.append({
            "name": s.name,
            "reg_number": s.reg_number,
            "stream_name": s.stream.name if s.stream else "—",
            "total_logs": tot,
            "present_logs": pres,
            "attendance_percentage": f"{percentage_val}%",
            "percentage_num": percentage_val
        })

    # 4. INTERCEPT SPECIFIC DETAILED EXPORT REQUESTS
    if request.GET.get('export') == 'detailed_student_csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="detailed_student_report_{today}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Registration Number', 'Assigned Stream', 'Total Sessions Checked', 'Sessions Attended', 'Attendance Rate'])
        for item in detailed_student_reports:
            writer.writerow([item['name'], item['reg_number'], item['stream_name'], item['total_logs'], item['present_logs'], item['attendance_percentage']])
        return response

    # 5. PRESERVE ORIGINAL ANALYTICS COMPONENT ARRAYS
    selected_stream_id = request.GET.get('stream')
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related('timetable_entry__teacher__user', 'timetable_entry__stream').order_by('-date_marked')
    
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # Process original global metrics
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # Location Log Parsing
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # Standard Top-20 Matrix Ingestion 
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # Lecturer Compliance Ingestion
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    total_system_sessions = sessions.count()
    teacher_counts = sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}
    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    
    if selected_stream_id and selected_stream_id != 'all':
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    teachers_raw = []
    seen_teachers = set()
    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            teachers_raw.append({'email': email, 'submitted': submitted, 'rate': f"{rate}%"})

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # Chart Generation Data Processing
    chart_dist_data = [present_count, absent_count]
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
        
        # New Detailed reporting fields injected safely into context context
        'departments': all_departments,
        'courses': selectable_courses,
        'selectable_streams': selectable_streams,
        'detailed_student_reports': detailed_student_reports,
        'filter_dept': filter_dept,
        'filter_course': filter_course,
        'filter_stream': filter_stream,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'quick_range': quick_range,
    }
    return render(request, 'attendance/analytics_dashboard.html', context)
    # Enforce strict Admin-only access rule matrix
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)

    # Extract optional filter parameters from template request scopes
    selected_stream_id = request.GET.get('stream')
    
    # 1. INITIALIZE VARIABLES FIRST (Fixes UnboundLocalError)
    # Base query sets targeting core transactional datasets
    records = AttendanceRecord.objects.all()
    sessions = AttendanceSession.objects.select_related(
        'timetable_entry__teacher__user', 
        'timetable_entry__stream'
    ).order_by('-date_marked')
    
    # 2. APPLY SCOPING FILTERS dynamically if requested
    if selected_stream_id and selected_stream_id != 'all':
        records = records.filter(student__stream_id=selected_stream_id)
        sessions = sessions.filter(timetable_entry__stream_id=selected_stream_id)

    # 3. INTERCEPT CSV EXPORT REQUEST
    # Now that 'records' exists and is filtered, we can safely export it!
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Name', 'Reg Number', 'Class/Stream', 'Status', 'Date'])
        
        # NOTE: If your relation to AttendanceSession is named something else 
        # (e.g. 'attendance_session'), change 'session' below to match your models.py
        for record in records.select_related('student__stream', 'session'):
            date_val = record.session.date_marked.strftime('%Y-%m-%d') if hasattr(record, 'session') and record.session.date_marked else 'N/A'
            stream_name = record.student.stream.name if record.student.stream else 'N/A'
            
            writer.writerow([
                record.student.name, 
                record.student.reg_number,
                stream_name,
                record.status, 
                date_val
            ])
        return response

    # 4. Compute Live Summary Statistics Metrics Matrix
    total_records = records.count()
    present_count = records.filter(status='PRESENT').count()
    absent_count = records.filter(status='ABSENT').count()

    stats = {
        'total_records': total_records,
        'present_rate': round((present_count / total_records) * 100, 1) if total_records > 0 else 0,
        'absent_rate': round((absent_count / total_records) * 100, 1) if total_records > 0 else 0,
    }

    # 5. Compile Lecturer Location Audit Log Dataset Dynamically
    audit_logs = []
    for session in sessions[:20]:
        t_entry = session.timetable_entry
        has_coords = session.teacher_latitude is not None and session.teacher_longitude is not None
        location_str = f"{session.teacher_latitude}, {session.teacher_longitude}" if has_coords else "Not captured"
        
        audit_logs.append({
            "date": session.date_marked.strftime("%Y-%m-%d") if session.date_marked else "—",
            "class": t_entry.stream.name if t_entry.stream else "—",
            "lecturer": t_entry.teacher.user.email if t_entry.teacher and t_entry.teacher.user else "—",
            "location": location_str,
            "ip": "—",  
            "verified": has_coords  
        })

    # 6. Pull Top Present & Absent Students Metrics Arrays via Database Annotations (Top 20)
    students = StudentProfile.objects.select_related('stream').annotate(
        total_rec=Count('attendancerecord'),
        present_rec=Count('attendancerecord', filter=Q(attendancerecord__status='PRESENT')),
        absent_rec=Count('attendancerecord', filter=Q(attendancerecord__status='ABSENT'))
    ).filter(total_rec__gt=0)
    
    students_raw = []
    for s in students:
        p_rate = (s.present_rec / s.total_rec) * 100
        a_rate = (s.absent_rec / s.total_rec) * 100
        students_raw.append({
            "name": s.name,
            "reg": s.reg_number,
            "class": s.stream.name if s.stream else "—",
            "p_rate_num": p_rate,
            "a_rate_num": a_rate,
            "p_rate": f"{round(p_rate, 1)}%",
            "a_rate": f"{round(a_rate, 1)}%"
        })
    
    top_present_students = sorted(students_raw, key=lambda x: x['p_rate_num'], reverse=True)[:20]
    top_absent_students = sorted(students_raw, key=lambda x: x['a_rate_num'], reverse=True)[:20]

    # 7. Generate Teacher Attendance Submission Compliance Metrics Report
    TimetableEntry = AttendanceSession._meta.get_field('timetable_entry').related_model
    total_system_sessions = sessions.count()

    teacher_counts = sessions.values('timetable_entry__teacher_id').annotate(count=Count('id'))
    counts_map = {item['timetable_entry__teacher_id']: item['count'] for item in teacher_counts if item['timetable_entry__teacher_id']}

    timetable_entries = TimetableEntry.objects.select_related('teacher__user').all()
    if selected_stream_id and selected_stream_id != 'all':
        timetable_entries = timetable_entries.filter(stream_id=selected_stream_id)

    teachers_raw = []
    seen_teachers = set()

    for entry in timetable_entries:
        if entry.teacher and entry.teacher.id not in seen_teachers:
            seen_teachers.add(entry.teacher.id)
            email = entry.teacher.user.email if entry.teacher.user else "—"
            submitted = counts_map.get(entry.teacher.id, 0)
            
            rate = round((submitted / total_system_sessions) * 100, 1) if total_system_sessions > 0 else 0
            
            teachers_raw.append({
                'email': email,
                'submitted': submitted,
                'rate': f"{rate}%"
            })

    top_submitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=True)[:20]
    top_unsubmitted_teachers = sorted(teachers_raw, key=lambda x: x['submitted'], reverse=False)[:20]

    # 8. Generate Aggregated Payload Engines for ChartJS
    chart_dist_data = [present_count, absent_count]
    
    streams_data = Stream.objects.annotate(
        total=Count('students__attendancerecord'),
        present=Count('students__attendancerecord', filter=Q(students__attendancerecord__status='PRESENT'))
    ).filter(total__gt=0)
    
    stream_labels = [stream.name for stream in streams_data]
    stream_rates = [round((stream.present / stream.total) * 100, 1) for stream in streams_data]

    context = {
        'stats': stats,
        'audit_logs': audit_logs,
        'top_present_students': top_present_students,
        'top_absent_students': top_absent_students,
        'top_submitted_teachers': top_submitted_teachers,
        'top_unsubmitted_teachers': top_unsubmitted_teachers,
        'streams': Stream.objects.all(),
        'selected_stream': selected_stream_id,
        'chart_dist_data': json.dumps(chart_dist_data),
        'stream_labels': json.dumps(stream_labels),
        'stream_rates': json.dumps(stream_rates),
    }
    return render(request, 'attendance/analytics_dashboard.html', context)



from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .models import TimetableBatch, TimetableEntry

@login_required
def export_timetable_pdf(request):
    """
    Generates a formal landscape PDF matrix tracking the current 
    active master schedule layout configurations.
    """
    # 1. Fetch the exact same active master batch as the editor view
    batch = TimetableBatch.objects.filter(is_active=True, is_revoked=False).order_by('-uploaded_at').first() #
    if not batch:
        messages.warning(request, "No active timetable configuration was found to generate a document.")
        return redirect('attendance:upload_timetable')
        
    # 2. Initialize HTTP Response object configured as PDF payload stream
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="timetable_batch_{batch.id}.pdf"' #
    
    # Setup document geometry in landscape format to accommodate the 7 day tracks
    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(letter), 
        rightMargin=30, 
        leftMargin=30, 
        topMargin=35, 
        bottomMargin=30
    )
    story = []
    
    # 3. Setup Layout Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor('#2b3e85'),
        alignment=1, # Centered alignment
        spaceAfter=15
    )
    
    cell_text_style = ParagraphStyle(
        'MatrixCell',
        parent=styles['Normal'],
        fontSize=7.5,
        leading=10,
        alignment=1 # Center text wrapped within table nodes
    )
    
    header_style = ParagraphStyle(
        'MatrixHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        fontName='Helvetica-Bold',
        textColor=colors.whitesmoke,
        alignment=1
    )
    
    # Document header markup text block
    formatted_date = batch.week_start_date.strftime('%B %d, %Y') #
    story.append(Paragraph(f"MASTER TIMETABLE SCHEDULE GRID", title_style))
    story.append(Paragraph(f"Active Schedule Target Cycle — Week Commencing: {formatted_date}", ParagraphStyle('Sub', alignment=1, fontSize=10, textColor=colors.HexColor('#475569'))))
    story.append(Spacer(1, 20))
    
    # 4. Extract and Group DB Matrix Entries (identical workflow as editor GET request)
    DAYS = [code for code, _ in TimetableEntry.DAYS_OF_WEEK] #
    headers = [Paragraph("Time Block Interval", header_style)] + [Paragraph(name, header_style) for _, name in TimetableEntry.DAYS_OF_WEEK] #
    
    entries = TimetableEntry.objects.filter(batch=batch).order_by('start_time') #
    grouped_slots = {}
    for entry in entries: #
        time_key = (entry.start_time.strftime('%H:%M'), entry.end_time.strftime('%H:%M')) #
        if time_key not in grouped_slots:
            grouped_slots[time_key] = {}
        grouped_slots[time_key][entry.day] = entry #
        
    matrix_data = [headers]
    
    # Build structural rows tracking each unique coordinate matrix node cell block
    for (start, end), day_map in grouped_slots.items():
        row_cells = [Paragraph(f"<b>{start} - {end}</b>", cell_text_style)]
        
        for day_code in DAYS:
            entry = day_map.get(day_code)
            if entry:
                # Text structural wrapping to safely output multi-line labels inside table cell nodes
                cell_content = f"<b>{entry.course_unit.code}</b><br/>{entry.teacher.name}<br/><font color='#475569'>{entry.stream.name}</font>" #
                row_cells.append(Paragraph(cell_content, cell_text_style))
            else:
                row_cells.append(Paragraph("<font color='#cbd5e1'>—</font>", cell_text_style))
                
        matrix_data.append(row_cells)
        
    # Calculate geometric column constraints to match page printable canvas space boundaries
    # Printable horizontal target width: 792 (landscape width) - 60 (margins) = 732 points total
    col_widths = [90] + [91] * 7 
    
    timetable_table = Table(matrix_data, colWidths=col_widths, repeatRows=1)
    timetable_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b3e85')), # Match original navy headers
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')), # Subtle structural slate boundaries
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]))
    
    story.append(timetable_table)
    
    # 5. Compile payload contents out to the client download pipeline stream
    doc.build(story)
    return response


@login_required
@transaction.atomic
def manage_users(request):
    # Guard to ensure only ADMIN can view or change roles
    if request.user.role != User.IS_ADMIN:
        return HttpResponse("Unauthorized", status=403)
        
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        new_role = request.POST.get('role')
        
        if user_id and new_role in dict(User.ROLE_CHOICES):
            target_user = get_object_or_404(User, id=user_id)
            target_user.role = new_role
            target_user.save()
            messages.success(request, f"Updated role for {target_user.username} to {new_role}.")
            return redirect('attendance:manage_users')

    all_users = User.objects.all().order_by('username')
    return render(request, 'attendance/manage_users.html', {
        'all_users': all_users,
        'role_choices': User.ROLE_CHOICES
    })